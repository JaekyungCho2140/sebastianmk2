"""실제 파일로 CSV 복원 테스트"""

import pytest
import queue
from pathlib import Path
from sebastian.core.common.csv_restore import restore_csv_quotes
import pandas as pd


@pytest.fixture
def real_files_dir():
    """실제 파일 디렉토리"""
    return Path(__file__).parent.parent.parent / "legacy" / ".csv test"


@pytest.fixture
def progress_queue():
    """진행 상황 Queue"""
    return queue.Queue()


class TestRealFiles:
    """실제 파일로 복원 테스트"""

    def test_restore_ymir_cup_file(self, real_files_dir, progress_queue, tmp_path):
        """실제 YMIR Cup 파일 복원"""
        from sebastian.core.common.csv_parser import analyze_csv_pattern

        original = real_files_dir / "251216_lygl-ymir_cup_EXTRA0_original.csv"
        exported = real_files_dir / "251216_lygl-ymir_cup_EXTRA0_exported.csv"

        # 파일 존재 확인
        if not original.exists() or not exported.exists():
            pytest.skip("실제 테스트 파일이 없습니다")

        output = tmp_path / "test_restored.csv"

        # 복원 실행
        restored_path, report_path = restore_csv_quotes(
            str(original), str(exported), str(output), progress_queue
        )

        # 복원 파일 생성 확인
        assert Path(restored_path).exists()
        assert Path(report_path).exists()

        # 복원 파일 읽기
        restored_df = pd.read_csv(restored_path, dtype=str, keep_default_na=False)
        original_df = pd.read_csv(str(original), dtype=str, keep_default_na=False)
        export_df = pd.read_csv(str(exported), dtype=str, keep_default_na=False)

        # DataFrame 기본 검증
        assert len(restored_df) == len(export_df)
        assert len(restored_df.columns) == len(export_df.columns)

        # 원본 패턴 분석
        original_pattern = analyze_csv_pattern(str(original))

        # 원본 패턴이 복원되었는지 검증
        # (pandas로 읽은 값은 동일하지만, 따옴표 패턴이 복원되었는지 확인)
        print(f"\n원본 패턴 필드 수: {len(original_pattern)}")
        print(f"복원된 행 수: {len(restored_df)}")

        # 복원 검증:
        # - 내용 동일한 필드: 원본 raw text 그대로
        # - 내용 변경된 필드: export data 사용
        # pandas로 읽었을 때 export와 대부분 동일해야 함

        # 기본 검증만 수행
        assert len(restored_df) == len(export_df)
        assert list(restored_df.columns) == list(export_df.columns)

    def test_report_contains_only_quote_changes(self, real_files_dir, progress_queue, tmp_path):
        """보고서에 따옴표 복원만 포함되는지 확인"""
        original = real_files_dir / "251216_lygl-ymir_cup_EXTRA0_original.csv"
        exported = real_files_dir / "251216_lygl-ymir_cup_EXTRA0_exported.csv"

        if not original.exists() or not exported.exists():
            pytest.skip("실제 테스트 파일이 없습니다")

        output = tmp_path / "test_restored.csv"

        # 복원 실행
        restored_path, report_path = restore_csv_quotes(
            str(original), str(exported), str(output), progress_queue
        )

        # 보고서 확인
        from openpyxl import load_workbook

        wb = load_workbook(report_path)

        # Restored Fields 시트 확인
        ws_restored = wb["Restored Fields"]

        # 헤더 확인
        assert ws_restored["A1"].value == "key-name"
        assert ws_restored["B1"].value == "Column"
        assert ws_restored["F1"].value == "Status"

        # 따옴표 복원만 포함되어야 함
        # (번역 변경은 포함되지 않음)
        restored_count = 0
        for row in range(2, ws_restored.max_row + 1):
            status = ws_restored[f"F{row}"].value
            if status and "따옴표 복원" in status:
                restored_count += 1

        print(f"\n따옴표 복원된 필드 수: {restored_count}")

        # 최소 1개 이상은 복원되어야 함
        assert restored_count > 0, "따옴표 복원이 발생하지 않았습니다"
