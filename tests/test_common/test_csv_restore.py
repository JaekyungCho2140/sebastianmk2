"""CSV Restore 테스트"""

import pytest
import queue
from pathlib import Path
from sebastian.core.common.csv_restore import restore_csv_quotes, generate_diff_report
import pandas as pd


@pytest.fixture
def sample_data_dir():
    """샘플 데이터 디렉토리 경로"""
    return Path(__file__).parent / "sample_data"


@pytest.fixture
def progress_queue():
    """진행 상황 Queue"""
    return queue.Queue()


class TestCSVRestore:
    """CSV 복원 테스트"""

    def test_restore_quotes_simple(self, sample_data_dir, progress_queue, tmp_path):
        """단순 따옴표 복원"""
        original = sample_data_dir / "original_normal.csv"
        export = sample_data_dir / "export_normal.csv"
        output = tmp_path / "restored.csv"

        restored_path, report_path = restore_csv_quotes(
            str(original), str(export), str(output), progress_queue
        )

        # 복원 파일 검증
        assert Path(restored_path).exists()
        assert restored_path.endswith("restored.csv")

        # 보고서 파일 검증
        assert Path(report_path).exists()
        assert report_path.endswith("_diff_report.xlsx")

        # 복원된 DataFrame 검증
        restored_df = pd.read_csv(restored_path, dtype=str, keep_default_na=False)
        assert len(restored_df) == 7  # 7개 행
        assert len(restored_df.columns) == 9  # 9개 컬럼

    def test_generate_diff_report(self, sample_data_dir, tmp_path):
        """차이점 보고서 생성 테스트"""
        from sebastian.core.common.csv_parser import analyze_csv_pattern, save_csv_with_pattern

        original_df = pd.read_csv(
            sample_data_dir / "original_normal.csv", dtype=str, keep_default_na=False
        )
        export_df = pd.read_csv(
            sample_data_dir / "export_normal.csv", dtype=str, keep_default_na=False
        )
        restored_df = export_df.copy()

        # 패턴 분석
        original_pattern = analyze_csv_pattern(str(sample_data_dir / "original_normal.csv"))
        export_pattern = analyze_csv_pattern(str(sample_data_dir / "export_normal.csv"))

        # 복원 패턴 (테스트용)
        restored_file = tmp_path / "test_restored.csv"
        restored_pattern = save_csv_with_pattern(
            restored_df, str(restored_file), original_pattern, original_df
        )

        report_path = tmp_path / "report.xlsx"

        result_path = generate_diff_report(
            original_df, export_df, restored_df, str(report_path),
            original_pattern, export_pattern, restored_pattern
        )

        # 보고서 파일 생성 확인
        assert Path(result_path).exists()
        assert result_path == str(report_path)

        # Excel 파일 구조 검증
        from openpyxl import load_workbook

        wb = load_workbook(result_path)
        assert "Summary" in wb.sheetnames
        assert "Restored Fields" in wb.sheetnames
        assert "Warnings" in wb.sheetnames

        # Summary 시트 검증
        ws_summary = wb["Summary"]
        assert ws_summary["A1"].value == "항목"
        assert ws_summary["B1"].value == "값"
        assert ws_summary["A2"].value == "총 행 수"

    def test_restore_with_validation_error(self, sample_data_dir, progress_queue, tmp_path):
        """검증 에러가 있는 경우"""
        from sebastian.core.common.csv_validator import CSVValidationError

        original = sample_data_dir / "error_column_mismatch_original.csv"
        export = sample_data_dir / "error_column_mismatch_export.csv"
        output = tmp_path / "restored.csv"

        with pytest.raises(CSVValidationError):
            restore_csv_quotes(str(original), str(export), str(output), progress_queue)

    def test_progress_updates(self, sample_data_dir, progress_queue, tmp_path):
        """진행 상황 업데이트 테스트"""
        original = sample_data_dir / "original_normal.csv"
        export = sample_data_dir / "export_normal.csv"
        output = tmp_path / "restored.csv"

        restored_path, report_path = restore_csv_quotes(
            str(original), str(export), str(output), progress_queue
        )

        # Queue에서 진행 상황 확인
        progress_messages = []
        status_messages = []

        while not progress_queue.empty():
            msg_type, msg_value = progress_queue.get()
            if msg_type == "progress":
                progress_messages.append(msg_value)
            elif msg_type == "status":
                status_messages.append(msg_value)

        # 진행률이 0~100 범위인지 확인
        assert all(0 <= p <= 100 for p in progress_messages)
        # 상태 메시지가 있는지 확인
        assert len(status_messages) > 0
        # 마지막 진행률이 100인지 확인
        assert 100 in progress_messages
