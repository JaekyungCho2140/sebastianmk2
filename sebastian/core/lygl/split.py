"""
분할 로직 모듈

PRD 섹션 2.2.2 "Split Operation"에 정의된 알고리즘을 구현합니다.
1개 통합 파일을 7개 언어별 파일로 분할합니다.
"""

from typing import Dict, Optional
from pathlib import Path
from openpyxl import load_workbook, Workbook

from .validator import (
    LANGUAGE_MAPPING,
    LANGUAGE_ORDER,
    ValidationError,
    validate_headers,
    validate_key,
    normalize_empty_value,
)
from .excel_format import apply_split_format


def split(merged_file_path: Path) -> Dict[str, Workbook]:
    """
    1개 병합 파일을 7개 언어별 파일로 분할

    Args:
        merged_file_path: 병합 파일 경로

    Returns:
        {'EN': Workbook, 'CT': Workbook, ...}

    Raises:
        ValidationError: 헤더 불일치, 빈 행 등
        FileNotFoundError: 파일이 존재하지 않을 시
        IOError: 파일 읽기 실패 시

    Reference:
        PRD 섹션 2.2.2 "Split Operation"
    """
    # 파일 존재 확인
    merged_path = Path(merged_file_path)
    if not merged_path.exists():
        raise FileNotFoundError(f"File not found: {merged_path}")

    # 1. 병합 파일 로드
    try:
        merged_wb = load_workbook(merged_path)
    except Exception as e:
        raise IOError(f"Failed to read merged file: {e}")

    merged_ws = merged_wb.active

    # 헤더 검증
    expected_headers = [
        "Table",
        "KEY",
        "Source",
        "Target_EN",
        "Target_CT",
        "Target_CS",
        "Target_JA",
        "Target_TH",
        "Target_PT",
        "Target_RU",
        "Status",
        "NOTE",
        "Date",
    ]
    actual_headers = [cell.value for cell in merged_ws[1]]
    validate_headers(actual_headers, expected_headers, merged_path.name)

    # 2. 각 언어별 파일 생성
    result_workbooks = {}

    for lang_code in LANGUAGE_ORDER:
        lang_wb = Workbook()
        lang_ws = lang_wb.active
        lang_ws.title = "Sheet1"  # 시트명을 'Sheet1'으로 설정 (MS Excel 기본값)

        # 헤더 작성
        lang_ws.append(["Table", "KEY", "Source", "Target", "Status", "NOTE", "Date"])

        # Target 컬럼 인덱스 계산
        target_column_name = LANGUAGE_MAPPING[lang_code]["column_name"]
        target_col_index = expected_headers.index(target_column_name)

        # 데이터 추출
        for idx, row in enumerate(
            merged_ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if len(row) < 13:
                # 행이 불완전한 경우 검증
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    # 일부 데이터가 있으면 에러
                    raise ValidationError(
                        f"Incomplete row found in merged file at row {idx}"
                    )
                continue  # 완전히 빈 행은 스킵

            table = row[0]
            key = row[1]
            source = row[2]
            target = row[target_col_index] if len(row) > target_col_index else ""
            status = row[10] if len(row) > 10 else ""
            note = row[11] if len(row) > 11 else ""
            date = row[12] if len(row) > 12 else ""

            # KEY 검증
            validate_key(key, idx, merged_path.name)

            # 값 정규화
            target = normalize_empty_value(target)
            status = normalize_empty_value(status) if status else ""
            note = normalize_empty_value(note)
            date = normalize_empty_value(date)

            lang_ws.append([table, key, source, target, status, note, date])

        # Split 전용 서식 적용 (헤더 배경색 없음, 틀 고정 없음)
        apply_split_format(lang_ws)

        result_workbooks[lang_code] = lang_wb

    return result_workbooks


def split_file(
    merged_file_path: str,
    output_directory: str,
    date_prefix: Optional[str] = None,
    progress_callback=None,
) -> Dict[str, str]:
    """
    병합 파일을 분할하여 디렉토리에 저장

    Args:
        merged_file_path: 병합 파일 경로
        output_directory: 출력 디렉토리 경로
        date_prefix: 파일명 날짜 접두사 (예: '251104'). None이면 입력 파일에서 추출
        progress_callback: 진행률 콜백 함수 (optional)

    Returns:
        {'EN': 'path/to/EN.xlsx', ...} 생성된 파일 경로 매핑

    Raises:
        ValidationError: 검증 실패 시
        FileNotFoundError: 파일이 존재하지 않을 시
        IOError: 파일 읽기/쓰기 실패 시
    """
    merged_path = Path(merged_file_path)
    output_dir = Path(output_directory)

    # 출력 디렉토리 생성
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise IOError(f"Failed to create output directory: {e}")

    if progress_callback:
        progress_callback(0, "분할 작업을 시작합니다...")

    # 날짜 추출 (입력 파일명에서)
    if not date_prefix:
        from .validator import extract_date

        date_prefix = extract_date(merged_path)
        if not date_prefix:
            # 날짜 추출 실패 시 현재 날짜 사용
            from datetime import datetime

            date_prefix = datetime.now().strftime("%y%m%d")

    # 분할 수행
    workbooks = split(merged_path)

    if progress_callback:
        progress_callback(50, "언어별 파일을 저장하는 중...")

    # 파일 저장
    output_paths = {}
    total_files = len(LANGUAGE_ORDER)

    for idx, lang_code in enumerate(LANGUAGE_ORDER):
        # 파일명 생성
        file_name = f"{date_prefix}_{LANGUAGE_MAPPING[lang_code]['file_name']}"
        output_path = output_dir / file_name

        # 저장
        try:
            workbooks[lang_code].save(output_path)
        except Exception as e:
            raise IOError(f"Failed to write {lang_code} file: {e}")

        output_paths[lang_code] = str(output_path)

        # 진행률 업데이트
        if progress_callback:
            progress = 50 + int((idx + 1) / total_files * 50)
            progress_callback(
                progress, f"{lang_code} 파일 저장 중 ({idx + 1}/{total_files})..."
            )

    if progress_callback:
        progress_callback(100, "분할이 완료되었습니다")

    return output_paths
