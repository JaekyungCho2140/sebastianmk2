"""
병합 로직 모듈

PRD 섹션 2.2.1 "Merge Operation"에 정의된 알고리즘을 구현합니다.
7개 언어별 파일을 1개 통합 파일로 병합합니다.
"""

from typing import Dict
from pathlib import Path
from openpyxl import load_workbook, Workbook

from .validator import (
    LANGUAGE_MAPPING,
    LANGUAGE_ORDER,
    ValidationError,
    validate_language_files,
    validate_headers,
    validate_key,
    validate_row_match,
    normalize_empty_value,
)
from .excel_format import apply_excel_format


def merge(language_files: Dict[str, Path]) -> Workbook:
    """
    7개 언어별 파일을 1개 병합 파일로 통합

    Args:
        language_files: {'EN': Path('path/to/EN.xlsx'), 'CT': Path(...), ...}

    Returns:
        병합된 Workbook 객체

    Raises:
        ValidationError: 파일 수 불일치, KEY 불일치, Table/Source 불일치 시
        FileNotFoundError: 파일이 존재하지 않을 시
        IOError: 파일 읽기 실패 시

    Reference:
        PRD 섹션 2.2.1 "Merge Operation"
    """
    # 0. 파일 경로를 Path 객체로 변환
    file_paths = {
        lang: Path(path) if not isinstance(path, Path) else path
        for lang, path in language_files.items()
    }

    # 파일 존재 확인
    for lang, path in file_paths.items():
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

    # 1. 파일 수 및 언어 검증 (정확히 7개)
    validate_language_files(list(file_paths.values()))

    # 2. EN 파일 (마스터) 로드
    en_path = file_paths.get("EN")
    if not en_path:
        raise ValidationError("EN (master) file is required")

    try:
        en_wb = load_workbook(en_path)
    except Exception as e:
        raise IOError(f"Failed to read EN file: {e}")

    en_ws = en_wb.active
    en_data = {}  # {KEY: {Table, Source, Status, NOTE, Target_EN}}

    # 헤더 검증 (대소문자 구분)
    expected_headers = ["Table", "KEY", "Source", "Target", "Status", "NOTE", "Date"]
    actual_headers = [cell.value for cell in en_ws[1]]
    validate_headers(actual_headers, expected_headers, en_path.name)

    # EN 데이터 수집 (2행부터)
    for idx, row in enumerate(en_ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 6:
            continue  # 빈 행 스킵

        table, key, source, target_en, status, note = row[:6]
        date = row[6] if len(row) > 6 else ""

        # KEY 검증
        validate_key(key, idx, en_path.name)

        # 중복 KEY 검증
        if key in en_data:
            raise ValidationError(f"Duplicate KEY in EN file: {key}")

        en_data[key] = {
            "Table": table,
            "Source": source,
            "Target_EN": normalize_empty_value(target_en),
            "Status": status,
            "NOTE": normalize_empty_value(note),
            "Date": normalize_empty_value(date),
        }

    # 3. 나머지 언어 파일 처리
    for lang_code in ["CT", "CS", "JA", "TH", "PT-BR", "RU"]:
        lang_path = file_paths.get(lang_code)
        if not lang_path:
            raise ValidationError(f"Missing language file: {lang_code}")

        try:
            lang_wb = load_workbook(lang_path)
        except Exception as e:
            raise IOError(f"Failed to read {lang_code} file: {e}")

        lang_ws = lang_wb.active

        # 헤더 검증
        lang_headers = [cell.value for cell in lang_ws[1]]
        validate_headers(lang_headers, expected_headers, lang_path.name)

        # 데이터 검증 및 병합
        for idx, row in enumerate(
            lang_ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if len(row) < 6:
                continue  # 빈 행 스킵

            table, key, source, target, status, note = row[:6]
            date = row[6] if len(row) > 6 else ""

            # KEY 존재 여부 확인 (EN에 없으면 에러)
            if key not in en_data:
                raise ValidationError(
                    f"KEY '{key}' in {lang_code} not found in EN (master) file"
                )

            # Table, Source, Status, NOTE, Date 일치 확인
            en_row = en_data[key]
            lang_row = {
                "Table": table,
                "Source": source,
                "Status": status,
                "NOTE": normalize_empty_value(note),
                "Date": normalize_empty_value(date),
            }

            validate_row_match(key, en_row, lang_row, lang_code)

            # Target 값 병합
            column_name = LANGUAGE_MAPPING[lang_code]["column_name"]
            en_data[key][column_name] = normalize_empty_value(target)

    # 4. 병합 파일 생성
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "Sheet1"  # 시트명을 'Sheet1'으로 설정 (MS Excel 기본값)

    # 헤더 작성
    merged_ws.append(
        [
            "Table",
            "KEY",
            "Source",
            "Target_EN",
            "Target_CT",
            "Target_CS",
            "Target_JA",
            "Target_TH",
            "Target_PT",
            "Target_RU",
            "Status",
            "NOTE",
            "Date",
        ]
    )

    # 데이터 작성 (EN 파일 순서 유지)
    for key, data in en_data.items():
        merged_ws.append(
            [
                data["Table"],
                key,
                data["Source"],
                data.get("Target_EN", ""),
                data.get("Target_CT", ""),
                data.get("Target_CS", ""),
                data.get("Target_JA", ""),
                data.get("Target_TH", ""),
                data.get("Target_PT", ""),
                data.get("Target_RU", ""),
                data["Status"],
                data["NOTE"],
                data["Date"],
            ]
        )

    # 5. Excel 서식 적용
    apply_excel_format(merged_ws, is_merged=True)

    return merged_wb


def merge_files(
    language_file_paths: Dict[str, str], output_path: str, progress_callback=None
) -> None:
    """
    7개 언어별 파일을 병합하여 파일로 저장

    Args:
        language_file_paths: {'EN': 'path/to/EN.xlsx', ...}
        output_path: 출력 파일 경로
        progress_callback: 진행률 콜백 함수 (optional)

    Raises:
        ValidationError: 검증 실패 시
        FileNotFoundError: 파일이 존재하지 않을 시
        IOError: 파일 읽기/쓰기 실패 시
    """
    # Path 객체로 변환
    file_paths = {lang: Path(path) for lang, path in language_file_paths.items()}

    if progress_callback:
        progress_callback(0, "병합 작업을 시작합니다...")

    # 병합 수행
    merged_wb = merge(file_paths)

    if progress_callback:
        progress_callback(80, "병합된 파일을 저장하는 중...")

    # 파일 저장
    output = Path(output_path)
    try:
        merged_wb.save(output)
    except Exception as e:
        raise IOError(f"Failed to write output file: {e}")

    if progress_callback:
        progress_callback(100, "병합이 완료되었습니다")
