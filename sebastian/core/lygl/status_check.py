"""
Status Check 모듈

7개 언어 파일의 Status 통일 여부를 검증합니다.
EN 파일을 기준으로 다른 6개 언어의 Status를 비교하여 불일치하는 키만 보고합니다.
"""

from pathlib import Path
from typing import Dict, List, Optional, Callable, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


# 지원 언어 목록
VALID_LANGUAGES = ['EN', 'CT', 'CS', 'JA', 'TH', 'PT-BR', 'RU']


class StatusCheckError(Exception):
    """Status Check 처리 오류"""
    def __init__(self, message: str, error_code: str = None):
        self.message = message
        self.error_code = error_code
        super().__init__(self.message)


def read_language_file(file_path: Path) -> Dict[str, str]:
    """
    언어 파일 읽기

    Args:
        file_path: LY Table Split 파일 경로

    Returns:
        {KEY: Status} 매핑

    파일 구조:
        | Table | KEY | Source | Target | Status |

    예외:
        - 파일이 없으면 FileNotFoundError
        - 파일 형식이 잘못되면 StatusCheckError
    """
    if not file_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        key_status_map = {}

        # 첫 행은 헤더, 2행부터 데이터
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 5:
                table, key, source, target, status = row[:5]

                # KEY와 Status가 모두 있는 경우만 수집
                if key and status:
                    key_status_map[key] = status

        wb.close()

        return key_status_map

    except Exception as e:
        raise StatusCheckError(f"파일 읽기 실패: {file_path}\n오류: {str(e)}")


def calculate_status_statistics(
    all_data: Dict[str, Dict[str, str]]
) -> Dict[str, Dict[str, int]]:
    """
    각 언어별 Status 통계 계산

    Args:
        all_data: {언어코드: {KEY: Status}}

    Returns:
        {
            'EN': {'기존': 10, '번역필요': 5, '수정': 2, '완료': 3},
            'CT': {'기존': 10, '번역필요': 5, '수정': 2, '완료': 3},
            ...
        }
    """
    statistics = {}

    for lang in VALID_LANGUAGES:
        stats = {'기존': 0, '번역필요': 0, '수정': 0, '완료': 0}

        for status in all_data[lang].values():
            if status in stats:
                stats[status] += 1

        statistics[lang] = stats

    return statistics


def check_status_consistency(
    files: Dict[str, Path],
    progress_callback: Optional[Callable[[int, str], None]] = None
) -> Tuple[List[Dict], Dict[str, Dict[str, int]]]:
    """
    Status 통일 여부 검증 + 통계 계산

    Args:
        files: {언어코드: Path} (7개 언어 파일)
        progress_callback: (percent, message) 진행 상황 콜백

    Returns:
        (inconsistencies, statistics)
        - inconsistencies: 불일치하는 키 목록
        - statistics: 각 언어별 Status 통계

    처리 규칙:
        1. EN 파일의 모든 KEY를 기준으로 수집
        2. 각 KEY에 대해 7개 언어의 Status 수집
        3. Status가 모두 동일하면 일치, 하나라도 다르면 불일치
        4. 불일치하는 KEY만 반환

    예외:
        - 7개 파일이 아니면 StatusCheckError
    """
    # 1. 파일 개수 검증
    if len(files) != 7:
        raise StatusCheckError(
            f"7개 언어 파일이 필요합니다. 현재: {len(files)}개"
        )

    # 2. 언어 코드 검증
    if set(files.keys()) != set(VALID_LANGUAGES):
        missing = set(VALID_LANGUAGES) - set(files.keys())
        extra = set(files.keys()) - set(VALID_LANGUAGES)
        msg = "언어 코드가 올바르지 않습니다.\n"
        if missing:
            msg += f"누락: {', '.join(missing)}\n"
        if extra:
            msg += f"불필요: {', '.join(extra)}\n"
        raise StatusCheckError(msg)

    if progress_callback:
        progress_callback(10, "파일 읽기 중...")

    # 3. 각 언어 파일 읽기
    all_data = {}
    for idx, (lang, file_path) in enumerate(files.items()):
        if progress_callback:
            percent = 10 + int((idx + 1) / 7 * 30)
            progress_callback(percent, f"{lang} 파일 읽기 중...")

        all_data[lang] = read_language_file(file_path)

    if progress_callback:
        progress_callback(40, "Status 비교 중...")

    # 4. 통계 계산
    statistics = calculate_status_statistics(all_data)

    # 5. EN 파일의 모든 KEY 수집 (기준)
    en_keys = set(all_data['EN'].keys())

    # 6. 각 KEY별 Status 비교
    inconsistencies = []
    total_keys = len(en_keys)

    for idx, key in enumerate(sorted(en_keys)):
        if progress_callback and idx % 100 == 0:
            percent = 40 + int((idx + 1) / total_keys * 50)
            progress_callback(percent, f"비교 중... ({idx + 1}/{total_keys})")

        # 각 언어에서 이 KEY의 Status 수집
        statuses = {}
        for lang in VALID_LANGUAGES:
            # 해당 언어 파일에 KEY가 있으면 Status, 없으면 'Missing'
            statuses[lang] = all_data[lang].get(key, 'Missing')

        # 모든 Status가 동일한지 확인
        unique_statuses = set(statuses.values())
        is_consistent = len(unique_statuses) == 1

        # 불일치하는 경우만 결과에 추가
        if not is_consistent:
            inconsistencies.append({
                'key': key,
                'statuses': statuses,
                'is_consistent': False
            })

    if progress_callback:
        progress_callback(90, "비교 완료")

    return inconsistencies, statistics


def create_status_check_output(
    inconsistencies: List[Dict],
    statistics: Dict[str, Dict[str, int]],
    output_path: Path,
    progress_callback: Optional[Callable[[int, str], None]] = None
) -> None:
    """
    Status Check 결과를 Excel로 출력

    Args:
        inconsistencies: check_status_consistency() 결과 - 불일치 키 목록
        statistics: calculate_status_statistics() 결과 - 언어별 통계
        output_path: 출력 파일 경로
        progress_callback: (percent, message) 진행 상황 콜백

    출력 형식:
        - Overview 시트만 생성 (언어별 시트 없음)
        - 상단: Summary 테이블 (언어별 Status 통계)
        - 하단: 불일치 키 목록
    """
    if progress_callback:
        progress_callback(92, "Excel 파일 생성 중...")

    # 1. Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"

    current_row = 1

    # 2. Summary 섹션
    # 2-1. Summary 타이틀
    ws.merge_cells(f'A{current_row}:I{current_row}')
    title_cell = ws.cell(current_row, 1)
    title_cell.value = "Status 통계 (각 언어별)"
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    # 2-2. Summary 헤더
    summary_headers = ['언어', '기존', '번역필요', '수정', '완료', '합계']
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(current_row, col_idx)
        cell.value = header
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="DBEEF4", end_color="DBEEF4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    # 2-3. Summary 데이터
    for lang in VALID_LANGUAGES:
        stats = statistics[lang]
        total = sum(stats.values())

        ws.cell(current_row, 1).value = lang
        ws.cell(current_row, 2).value = stats['기존']
        ws.cell(current_row, 3).value = stats['번역필요']
        ws.cell(current_row, 4).value = stats['수정']
        ws.cell(current_row, 5).value = stats['완료']
        ws.cell(current_row, 6).value = total

        # 중앙 정렬
        for col_idx in range(1, 7):
            ws.cell(current_row, col_idx).alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1

    # 2-4. 빈 행 추가
    current_row += 2

    # 3. 불일치 키 목록 섹션
    # 3-1. 섹션 타이틀
    ws.merge_cells(f'A{current_row}:I{current_row}')
    section_cell = ws.cell(current_row, 1)
    section_cell.value = f"Status 불일치 키 목록 ({len(inconsistencies)}개)"
    section_cell.font = Font(name="Calibri", size=14, bold=True)
    section_cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    section_cell.alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    # 3-2. 헤더
    headers = ['#', 'KEY'] + VALID_LANGUAGES
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx)
        cell.value = header
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="DBEEF4", end_color="DBEEF4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    header_row = current_row
    current_row += 1

    if progress_callback:
        progress_callback(94, "데이터 작성 중...")

    # 3-3. 데이터 작성
    for idx, item in enumerate(inconsistencies, start=1):
        row_data = [idx, item['key']]

        # 각 언어의 Status 추가
        for lang in VALID_LANGUAGES:
            row_data.append(item['statuses'][lang])

        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(current_row, col_idx).value = value

        current_row += 1

    # 4. 데이터 스타일 적용
    if progress_callback:
        progress_callback(96, "스타일 적용 중...")

    # EN의 Status를 기준값으로 사용
    yellow_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    data_start_row = header_row + 1
    data_end_row = header_row + len(inconsistencies)

    for row_idx in range(data_start_row, data_end_row + 1):
        # EN Status (기준)
        en_status = ws.cell(row_idx, 3).value  # C열 (EN)

        # 각 언어 셀에 스타일 적용
        for col_idx in range(3, 10):  # C~I열 (EN~RU)
            cell = ws.cell(row_idx, col_idx)
            cell.alignment = center_align

            # Missing이면 빨간색
            if cell.value == 'Missing':
                cell.fill = red_fill
            # EN과 다르면 노란색
            elif cell.value != en_status:
                cell.fill = yellow_fill

    # 5. 컬럼 너비 설정
    ws.column_dimensions['A'].width = 6   # # 또는 언어
    ws.column_dimensions['B'].width = 50  # KEY 또는 기존
    ws.column_dimensions['C'].width = 12  # EN 또는 번역필요
    ws.column_dimensions['D'].width = 12  # CT 또는 수정
    ws.column_dimensions['E'].width = 12  # CS 또는 완료
    ws.column_dimensions['F'].width = 12  # JA 또는 합계
    ws.column_dimensions['G'].width = 12  # TH
    ws.column_dimensions['H'].width = 14  # PT-BR
    ws.column_dimensions['I'].width = 12  # RU

    # 6. 자동 필터 (불일치 키 목록에만)
    if len(inconsistencies) > 0:
        ws.auto_filter.ref = f"A{header_row}:I{data_end_row}"

    # 7. 파일 저장
    if progress_callback:
        progress_callback(98, "파일 저장 중...")

    wb.save(output_path)
    wb.close()

    if progress_callback:
        progress_callback(100, "완료")


def status_check(
    files: Dict[str, Path],
    output_path: Path,
    progress_callback: Optional[Callable[[int, str], None]] = None
) -> int:
    """
    Status Check 전체 프로세스 실행

    Args:
        files: {언어코드: Path}
        output_path: 출력 파일 경로
        progress_callback: (percent, message) 진행 상황 콜백

    Returns:
        불일치하는 키의 개수

    예외:
        StatusCheckError: 처리 중 오류 발생
    """
    # Step 1: Status 비교 + 통계 계산
    inconsistencies, statistics = check_status_consistency(files, progress_callback)

    # Step 2: 결과 출력 (통계 포함)
    create_status_check_output(inconsistencies, statistics, output_path, progress_callback)

    return len(inconsistencies)
