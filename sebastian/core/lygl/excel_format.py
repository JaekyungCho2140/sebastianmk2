"""
Excel 서식 적용 모듈

PRD 섹션 2.1.4 "Excel Format Specification"에 정의된 서식을 적용합니다.
"""

from typing import Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def apply_excel_format(worksheet: Worksheet, is_merged: bool = True) -> None:
    """
    Excel 워크시트에 표준 서식 적용

    Args:
        worksheet: openpyxl Worksheet 객체
        is_merged: True면 병합 파일 형식(12컬럼), False면 언어별 파일 형식(6컬럼)

    Reference:
        PRD 섹션 2.1.4 "Excel Format Specification"
    """
    # 1. 열 너비 설정
    if is_merged:
        # 병합 파일: 13개 컬럼
        widths = {
            "A": 25.71,  # Table
            "B": 40.71,  # KEY
            "C": 20.00,  # Source
            "D": 20.00,  # Target_EN
            "E": 20.00,  # Target_CT
            "F": 20.00,  # Target_CS
            "G": 20.00,  # Target_JA
            "H": 20.00,  # Target_TH
            "I": 20.00,  # Target_PT
            "J": 20.00,  # Target_RU
            "K": 12.71,  # Status
            "L": 30.71,  # NOTE
            "M": 20.00,  # Date
        }
    else:
        # 언어별 파일: 7개 컬럼
        widths = {
            "A": 25.71,  # Table
            "B": 40.71,  # KEY
            "C": 20.00,  # Source
            "D": 20.00,  # Target
            "E": 12.71,  # Status
            "F": 30.71,  # NOTE
            "G": 20.00,  # Date
        }

    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width

    # 2. 행 높이 설정
    if worksheet.max_row > 0:
        worksheet.row_dimensions[1].height = 16.5  # 헤더
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 30  # 데이터

    # 3. 헤더 서식 (1행)
    header_fill = PatternFill(
        start_color="DBEEF4", end_color="DBEEF4", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, color="000000")
    header_alignment = Alignment(
        horizontal="general", vertical="center", wrap_text=True
    )

    num_cols = 13 if is_merged else 7
    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 4. 데이터 서식 (2행 이상)
    data_font = Font(name="Calibri", size=11, color="000000")
    data_alignment = Alignment(horizontal="general", vertical="center", wrap_text=True)

    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, num_cols + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment

    # 5. 고정 틀 (첫 행 고정)
    worksheet.freeze_panes = "A2"

    # 6. 자동 필터
    if worksheet.max_row > 0:
        last_col = get_column_letter(num_cols)
        last_row = worksheet.max_row
        worksheet.auto_filter.ref = f"A1:{last_col}{last_row}"


def apply_split_format(worksheet: Worksheet) -> None:
    """
    Split 기능 전용 Excel 서식 적용

    원본 파일(.claude/docs/Tables by Language/)과 동일한 서식 적용:
    - 열 너비: 원본과 동일 (C=60.71, D=13.0)
    - 행 높이: 헤더는 기본값, 데이터는 30pt
    - 폰트: Calibri 11pt
    - 정렬: 수평은 기본값(None), 수직은 center, 줄바꿈 True
    - 배경색: 없음
    - 틀 고정: 없음
    - 자동 필터: 있음

    Args:
        worksheet: openpyxl Worksheet 객체

    Reference:
        원본 파일: .claude/docs/Tables by Language/251104_EN.xlsx
    """
    # 1. 열 너비 설정 (MS Excel 기준 원본 파일과 동일)
    widths = {
        "A": 25,  # Table
        "B": 40,  # KEY
        "C": 60,  # Source
        "D": 60,  # Target (MS Excel 확인: 60)
        "E": 12,  # Status
        "F": 30,  # NOTE
        "G": 20,  # Date
    }

    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width

    # 2. 행 높이 설정 (원본과 동일)
    if worksheet.max_row > 0:
        # 헤더는 기본값(None) 유지 - 명시적으로 설정하지 않음
        # worksheet.row_dimensions[1].height = None (기본값)

        # 데이터 행만 30pt 설정
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 30

    # 3. 폰트 및 정렬 (원본과 동일)
    font = Font(name="Calibri", size=11, color="000000")
    # 수평 정렬은 None(기본값), 수직 정렬은 center, 줄바꿈 True
    alignment = Alignment(horizontal=None, vertical="center", wrap_text=True)

    # 모든 셀에 폰트와 정렬 적용
    for row_idx in range(1, worksheet.max_row + 1):
        for col_idx in range(1, 8):  # 7개 컬럼
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = font
            cell.alignment = alignment

    # 4. 틀 고정 없음 (원본과 동일)
    # worksheet.freeze_panes = None (기본값)

    # 5. 자동 필터 적용 (원본과 동일)
    if worksheet.max_row > 0:
        last_row = worksheet.max_row
        worksheet.auto_filter.ref = f"A1:G{last_row}"


def get_column_widths(is_merged: bool = True) -> dict:
    """
    컬럼 너비 매핑 반환

    Args:
        is_merged: True면 병합 파일 형식, False면 언어별 파일 형식

    Returns:
        컬럼명 → 너비 매핑 딕셔너리
    """
    if is_merged:
        return {
            "A": 25.71,
            "B": 40.71,
            "C": 20.00,
            "D": 20.00,
            "E": 20.00,
            "F": 20.00,
            "G": 20.00,
            "H": 20.00,
            "I": 20.00,
            "J": 20.00,
            "K": 12.71,
            "L": 30.71,
            "M": 20.00,
        }
    else:
        return {"A": 25.71, "B": 40.71, "C": 20.00, "D": 20.00, "E": 12.71, "F": 30.71, "G": 20.00}
