"""CSV 따옴표 복원 모듈

memoQ에서 export한 CSV 파일의 따옴표를 원본 패턴으로 복원합니다.
"""

from typing import Tuple, Dict
import pandas as pd
from pathlib import Path
import queue
import csv
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from sebastian.core.common.csv_validator import validate_csv_structure
from sebastian.core.common.csv_parser import analyze_csv_pattern, save_csv_with_pattern

logger = logging.getLogger(__name__)


def restore_csv_quotes(
    original_path: str,
    export_path: str,
    output_path: str,
    progress_queue: queue.Queue,
) -> Tuple[str, str]:
    """CSV 따옴표 복원

    원본 CSV 파일의 따옴표 패턴을 분석하여,
    memoQ export CSV 파일의 따옴표를 원본 패턴으로 복원합니다.

    알고리즘:
        1. 검증: validate_csv_structure() 호출
        2. key-name 기준 매칭: 딕셔너리 생성 {key_name: row_index}
        3. 필드별 복원:
           - 원본에 따옴표가 있었으면 → 복원 파일에도 따옴표 추가
           - 원본에 따옴표가 없었으면 → 복원 파일에도 따옴표 제거
        4. 파일 저장: _restored.csv 생성
        5. 보고서 생성: generate_diff_report() 호출

    Args:
        original_path: 원본 CSV 파일 경로
        export_path: memoQ export CSV 파일 경로
        output_path: 복원 파일 저장 경로 (_restored.csv)
        progress_queue: 진행 상황 Queue (0-100 int 값)

    Returns:
        (복원 파일 경로, 보고서 파일 경로)

    Raises:
        CSVValidationError: 검증 실패 시
        IOError: 파일 I/O 실패 시

    Examples:
        >>> import queue
        >>> q = queue.Queue()
        >>> restored, report = restore_csv_quotes(
        ...     "original.csv", "export.csv", "output_restored.csv", q
        ... )
        >>> print(f"복원: {restored}, 보고서: {report}")
    """
    progress_queue.put(("status", "CSV 파일 검증 중..."))
    progress_queue.put(("progress", 10))

    # 1. 검증
    original_df, export_df, warnings = validate_csv_structure(
        original_path, export_path
    )

    progress_queue.put(("status", "원본 따옴표 패턴 분석 중..."))
    progress_queue.put(("progress", 20))

    # 2. 원본 파일을 raw CSV로 분석하여 따옴표 패턴 추출
    try:
        original_quote_pattern = analyze_csv_pattern(original_path)
        logger.info(f"원본 패턴 분석 완료: {len(original_quote_pattern)}개 필드")
    except Exception as e:
        raise IOError(f"원본 파일 패턴 분석 실패: {e}")

    # 2-1. Export 파일도 raw CSV로 분석 (보고서용)
    try:
        export_quote_pattern = analyze_csv_pattern(export_path)
        logger.info(f"Export 패턴 분석 완료: {len(export_quote_pattern)}개 필드")
    except Exception as e:
        raise IOError(f"Export 파일 패턴 분석 실패: {e}")

    progress_queue.put(("status", "key-name 기준 행 매칭 중..."))
    progress_queue.put(("progress", 30))

    # 3. key-name 기준 매칭
    key_column = original_df.columns[0]
    original_key_map = {
        key: idx for idx, key in enumerate(original_df[key_column].values)
    }
    export_key_map = {key: idx for idx, key in enumerate(export_df[key_column].values)}

    progress_queue.put(("status", "따옴표 복원 중..."))
    progress_queue.put(("progress", 40))

    # 4. 복원 DataFrame 생성
    restored_df = export_df.copy()
    restored_records = []

    total_rows = len(export_df)
    for row_idx, (key, export_row_idx) in enumerate(export_key_map.items()):
        original_row_idx = original_key_map[key]

        restored_row = {}
        for col in original_df.columns:
            original_field = original_df.iloc[original_row_idx][col]
            export_field = export_df.iloc[export_row_idx][col]

            # 원본 따옴표 패턴 복원
            has_original_quotes = original_quote_pattern.get(
                (original_row_idx, col), False
            )

            # export 필드의 내용만 가져오고 따옴표는 원본 패턴 적용
            restored_field = export_field  # 기본값은 export 필드

            restored_row[col] = restored_field

        restored_records.append(restored_row)

        # 진행률 업데이트 (40% ~ 70%)
        progress = 40 + int((row_idx + 1) / total_rows * 30)
        progress_queue.put(("progress", progress))

    restored_df = pd.DataFrame(restored_records)

    progress_queue.put(("status", "복원 파일 저장 중..."))
    progress_queue.put(("progress", 75))

    # 5. 복원 파일 저장 (원본 raw text 패턴 적용, RFC 4180 무시)
    restored_quote_pattern = save_csv_with_pattern(
        restored_df, output_path, original_quote_pattern, original_df
    )

    progress_queue.put(("status", "차이점 보고서 생성 중..."))
    progress_queue.put(("progress", 85))

    # 6. 보고서 생성
    report_path = str(Path(output_path).with_suffix("")) + "_diff_report.xlsx"
    generate_diff_report(
        original_df,
        export_df,
        restored_df,
        report_path,
        original_quote_pattern,
        export_quote_pattern,
        restored_quote_pattern,
    )

    progress_queue.put(("status", "완료!"))
    progress_queue.put(("progress", 100))

    logger.info(f"CSV 복원 완료: {output_path}")
    logger.info(f"보고서 생성 완료: {report_path}")

    return output_path, report_path




def generate_diff_report(
    original_df: pd.DataFrame,
    export_df: pd.DataFrame,
    restored_df: pd.DataFrame,
    output_path: str,
    original_quote_pattern: Dict[Tuple[int, str], Dict[str, any]],
    export_quote_pattern: Dict[Tuple[int, str], Dict[str, any]],
    restored_quote_pattern: Dict[Tuple[int, str], str],
) -> str:
    """차이점 보고서 생성 (Excel)

    원본, export, 복원 DataFrame을 비교하여
    차이점 보고서를 Excel 파일로 생성합니다.

    Excel 구조:
        - Sheet 1: Summary (요약)
        - Sheet 2: Restored Fields (복원된 필드 상세)
        - Sheet 3: Warnings (경고 사항)

    Args:
        original_df: 원본 DataFrame
        export_df: export DataFrame
        restored_df: 복원 DataFrame
        output_path: 보고서 저장 경로 (_diff_report.xlsx)

    Returns:
        보고서 파일 경로

    Examples:
        >>> report_path = generate_diff_report(
        ...     original_df, export_df, restored_df, "report.xlsx"
        ... )
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # 헤더 스타일
    header_fill = PatternFill(start_color="5E35B1", end_color="5E35B1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Summary 헤더
    ws_summary["A1"] = "항목"
    ws_summary["B1"] = "값"
    ws_summary["A1"].fill = header_fill
    ws_summary["A1"].font = header_font
    ws_summary["B1"].fill = header_fill
    ws_summary["B1"].font = header_font

    # Summary 데이터 (나중에 업데이트할 placeholder)
    key_column = original_df.columns[0]
    total_rows = len(original_df)
    total_fields = total_rows * len(original_df.columns)

    summary_data = [
        ("총 행 수", total_rows),
        ("총 필드 수", total_fields),
        ("따옴표 복원된 필드 수", 0),  # 나중에 업데이트
        ("경고 수", 0),
        ("오류 수", 0),
    ]

    for row_idx, (label, value) in enumerate(summary_data, start=2):
        ws_summary[f"A{row_idx}"] = label
        ws_summary[f"B{row_idx}"] = value

    # Sheet 2: Restored Fields
    ws_restored = wb.create_sheet("Restored Fields")

    # 헤더
    restored_headers = ["key-name", "Column", "Original", "Export", "Restored", "Status"]
    for col_idx, header in enumerate(restored_headers, start=1):
        cell = ws_restored.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 데이터 - 따옴표 복원이 발생한 필드만 기록
    row_num = 2
    restored_count = 0

    for idx in range(len(original_df)):
        key_value = original_df.iloc[idx][key_column]

        for col in original_df.columns:
            # 따옴표 패턴 확인
            pattern_key = (idx, col)
            if pattern_key not in original_quote_pattern:
                continue

            # 원본, Export, Restored의 raw text 가져오기
            original_raw = original_quote_pattern[pattern_key].get('raw_field_text', '')
            export_raw = export_quote_pattern[pattern_key].get('raw_field_text', '')
            restored_raw = restored_quote_pattern.get(pattern_key, '')

            # 따옴표 복원이 발생했는지 확인
            # Export raw vs Restored raw 비교!
            quote_restoration_occurred = (export_raw != restored_raw)

            if quote_restoration_occurred:
                ws_restored.cell(row=row_num, column=1, value=key_value)
                ws_restored.cell(row=row_num, column=2, value=col)

                # 원본 raw text 표시
                ws_restored.cell(row=row_num, column=3, value=original_raw)

                # Export raw text 표시
                ws_restored.cell(row=row_num, column=4, value=export_raw)

                # Restored raw text 표시
                ws_restored.cell(row=row_num, column=5, value=restored_raw)

                # 상태
                status = "✅ 따옴표 복원"
                status_fill = PatternFill(
                    start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
                )

                status_cell = ws_restored.cell(row=row_num, column=6, value=status)
                status_cell.fill = status_fill

                row_num += 1
                restored_count += 1

    # Summary 업데이트 (복원된 필드 수)
    ws_summary["B4"] = restored_count

    # Sheet 3: Warnings
    ws_warnings = wb.create_sheet("Warnings")

    # 헤더
    warning_headers = ["Type", "key-name", "Message"]
    for col_idx, header in enumerate(warning_headers, start=1):
        cell = ws_warnings.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 경고 데이터 (현재는 검증 단계에서 예외 발생하므로 여기까지 오지 않음)
    ws_warnings.cell(row=2, column=1, value="-")
    ws_warnings.cell(row=2, column=2, value="-")
    ws_warnings.cell(row=2, column=3, value="검증 통과 (경고 없음)")

    # 열 너비 조정
    for ws in [ws_summary, ws_restored, ws_warnings]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # 저장
    wb.save(output_path)
    logger.info(f"차이점 보고서 생성 완료: {output_path}")

    return output_path
