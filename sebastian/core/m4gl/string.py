import pandas as pd
import os
import datetime
import time
import stat
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


def profile_function(func):
    """함수 실행 시간 측정 데코레이터"""
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        print(f"{func.__name__} 실행 시간: {end - start:.2f}초")
        return result
    return wrapper


@profile_function
def read_excel_file(file_path, sheet_name, header_row, skip_rows):
    start_time = time.time()
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, skiprows=skip_rows)
    end_time = time.time()
    print(f"파일 읽기 시간 ({os.path.basename(file_path)}): {end_time - start_time:.2f}초")
    return df


def merge_string(folder_path: str, progress_queue) -> None:
    start_time = time.time()
    try:
        # 파일 경로 설정
        file_list = [
            "SEQUENCE_DIALOGUE.xlsm",
            "STRING_BUILTIN.xlsm",
            "STRING_MAIL.xlsm",
            "STRING_MESSAGE.xlsm",
            "STRING_NPC.xlsm",
            "STRING_QUESTTEMPLATE.xlsm",
            "STRING_TEMPLATE.xlsm",
            "STRING_TOOLTIP.xlsm"
        ]

        # 각 파일의 헤더와 데이터 시작 행
        header_rows = {
            "SEQUENCE_DIALOGUE.xlsm": 2,
            "STRING_BUILTIN.xlsm": 2,
            "STRING_MAIL.xlsm": 2,
            "STRING_MESSAGE.xlsm": 2,
            "STRING_NPC.xlsm": 2,
            "STRING_QUESTTEMPLATE.xlsm": 2,
            "STRING_TEMPLATE.xlsm": 2,
            "STRING_TOOLTIP.xlsm": 2
        }

        start_rows = {
            "SEQUENCE_DIALOGUE.xlsm": 9,
            "STRING_BUILTIN.xlsm": 4,
            "STRING_MAIL.xlsm": 4,
            "STRING_MESSAGE.xlsm": 4,
            "STRING_NPC.xlsm": 4,
            "STRING_QUESTTEMPLATE.xlsm": 7,
            "STRING_TEMPLATE.xlsm": 4,
            "STRING_TOOLTIP.xlsm": 4
        }

        # 매칭되는 열 인덱스 설정
        matching_columns = {
            "SEQUENCE_DIALOGUE.xlsm": [7, None, 10, 11, 12, 13, 14, 15, 16, 17, None, None],
            "STRING_BUILTIN.xlsm": [7, 21, 8, 9, 10, 11, 12, 13, 14, 15, None, None],
            "STRING_MAIL.xlsm": [7, None, 8, 9, 10, 11, 12, 13, 14, 15, None, None],
            "STRING_MESSAGE.xlsm": [7, 21, 8, 9, 10, 11, 12, 13, 14, 15, None, None],
            "STRING_NPC.xlsm": [7, 20, 9, 10, 11, 12, 13, 14, 15, 16, 18, 19],
            "STRING_QUESTTEMPLATE.xlsm": [7, 0, 12, 13, 14, 15, 16, 17, 18, 19, None, None],
            "STRING_TEMPLATE.xlsm": [7, 19, 8, 9, 10, 11, 12, 13, 14, 15, None, 18],
            "STRING_TOOLTIP.xlsm": [7, 8, 11, 12, 13, 14, 15, 16, 17, 18, None, None]
        }

        # 결과물 파일의 헤더 설정
        headers = ['#', 'Table Name', 'String ID', 'Table/ID', 'NOTE', 'KO', 'EN', 'CT', 'CS', 'JA', 'TH', 'ES-LATAM', 'PT-BR', 'NPC 이름', '비고']

        # 결과 데이터 프레임 생성
        result_df = pd.DataFrame(columns=headers)

        # 단계 정보 전송
        progress_queue.put("단계:1/2")
        progress_queue.put("파일:파일 읽는 중...")

        # 각 파일에서 데이터 읽어오기
        for i, file in enumerate(file_list):
            file_path = os.path.join(folder_path, file)
            if not os.path.isfile(file_path):
                progress_queue.put(("error", f"파일을 찾을 수 없습니다: {file_path}"))
                return

            progress_queue.put(f"파일:{file}")
            data = read_excel_file(file_path, sheet_name=1, header_row=header_rows[file], skip_rows=start_rows[file])
            # 글로벌 OnOFF=1 필터링 (G열, 인덱스 6)
            data = data[data.iloc[:, 6] == 1]

            # Table Name 열 채우기
            table_name = file.replace(".xlsm", "")
            temp_df = pd.DataFrame({
                '#': range(len(result_df) + 1, len(result_df) + len(data) + 1),
                'Table Name': table_name,
                'String ID': data.iloc[:, matching_columns[file][0]] if matching_columns[file][0] is not None else '',
                'Table/ID': table_name + '/' + data.iloc[:, matching_columns[file][0]].astype(str) if matching_columns[file][0] is not None else '',
                'NOTE': data.iloc[:, matching_columns[file][1]] if matching_columns[file][1] is not None else '',
                'KO': data.iloc[:, matching_columns[file][2]] if matching_columns[file][2] is not None else '',
                'EN': data.iloc[:, matching_columns[file][3]] if matching_columns[file][3] is not None else '',
                'CT': data.iloc[:, matching_columns[file][4]] if matching_columns[file][4] is not None else '',
                'CS': data.iloc[:, matching_columns[file][5]] if matching_columns[file][5] is not None else '',
                'JA': data.iloc[:, matching_columns[file][6]] if matching_columns[file][6] is not None else '',
                'TH': data.iloc[:, matching_columns[file][7]] if matching_columns[file][7] is not None else '',
                'ES-LATAM': data.iloc[:, matching_columns[file][8]] if matching_columns[file][8] is not None else '',
                'PT-BR': data.iloc[:, matching_columns[file][9]] if matching_columns[file][9] is not None else '',
                'NPC 이름': data.iloc[:, matching_columns[file][10]] if matching_columns[file][10] is not None else '',
                '비고': data.iloc[:, matching_columns[file][11]] if matching_columns[file][11] is not None else ''
            })
            result_df = pd.concat([result_df, temp_df], ignore_index=True)

            progress_queue.put(int(20 + (50 / len(file_list)) * (i + 1)))
            progress_queue.put(f"처리된 파일:{i+1}")

        progress_queue.put("단계:2/2")
        progress_queue.put("파일:결과 파일 저장 중...")

        # String ID를 정수로 변환 (소수점 제거)
        result_df['String ID'] = pd.to_numeric(result_df['String ID'], errors='coerce').fillna(0).astype('int64')
        # Table/ID 재생성 (정수 기반)
        result_df['Table/ID'] = result_df['Table Name'] + '/' + result_df['String ID'].astype(str)

        # 7번째 열(인덱스 6)이 빈 셀(NaN) 또는 0 또는 '미사용'인 행 제거
        result_df = result_df[~(pd.isna(result_df.iloc[:, 6]) | result_df.iloc[:, 6].isin([0, '미사용']))]

        # 인덱스 열 갱신
        result_df['#'] = range(1, len(result_df) + 1)

        # 출력 파일 이름 설정
        date_str = datetime.datetime.now().strftime('%m%d')
        output_file = f'{date_str}_MIR4_MASTER_STRING.xlsx'

        # 파일이 이미 존재할 경우 다른 이름으로 저장
        counter = 1
        while os.path.exists(output_file):
            output_file = f'{date_str}_MIR4_MASTER_STRING_{counter}.xlsx'
            counter += 1

        # 결과 파일 저장 (엑셀)
        result_df.to_excel(output_file, index=False)

        # 결과 파일 서식 지정
        wb = load_workbook(output_file)
        ws = wb.active

        # 폰트 및 서식 설정
        header_font = Font(name='맑은 고딕', size=12, bold=True, color='9C5700')
        default_font = Font(name='맑은 고딕', size=10)
        header_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        border_style = Side(border_style='thin', color='000000')
        full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        # 헤더 행 서식 지정
        if isinstance(ws, Worksheet):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = full_border

            # 나머지 셀 서식 지정
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = default_font
                    cell.border = full_border

            # 틀 고정
            ws.freeze_panes = 'A2'

        # 서식 지정된 파일 저장
        wb.save(output_file)

        # 결과 파일 읽기 전용 설정
        os.chmod(output_file, stat.S_IREAD)

        progress_queue.put(100)

        elapsed_time = time.time() - start_time
        progress_queue.put(f"완료:파일이 {output_file}로 저장되었습니다. 소요 시간: {int(elapsed_time)}초")

    except Exception as e:
        progress_queue.put(("error", str(e)))
