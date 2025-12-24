import pandas as pd
import os
import datetime
import time
import stat
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


def profile_function(func):
    """함수 실행 시간 측정 데코레이터"""
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        print(f"{func.__name__} 실행 시간: {end - start:.2f}초")
        return result
    return wrapper


@profile_function
def read_excel_file(file_path, sheet_name, header_row, skip_rows):
    start_time = time.time()
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, skiprows=skip_rows)
    end_time = time.time()
    print(f"파일 읽기 시간 ({os.path.basename(file_path)}): {end_time - start_time:.2f}초")
    return df


# 데이터 읽기 전에 먼저 파일이 존재하는지 확인하고, 열이 존재하는지 확인
def merge_dialogue(folder_path: str, progress_queue) -> None:
    start_time = time.time()
    try:
        # 파일 경로 설정
        cinematic_path = os.path.join(folder_path, "CINEMATIC_DIALOGUE.xlsm")
        smalltalk_path = os.path.join(folder_path, "SMALLTALK_DIALOGUE.xlsm")
        npc_path = os.path.join(folder_path, "NPC.xlsm")

        # 파일 존재 여부 확인
        missing_files = []
        if not os.path.isfile(cinematic_path):
            missing_files.append(f"파일을 찾을 수 없습니다: {cinematic_path}")
        if not os.path.isfile(smalltalk_path):
            missing_files.append(f"파일을 찾을 수 없습니다: {smalltalk_path}")
        if not os.path.isfile(npc_path):
            missing_files.append(f"파일을 찾을 수 없습니다: {npc_path}")
        if missing_files:
            raise FileNotFoundError("\n".join(missing_files))

        # 단계 정보 전송
        progress_queue.put("단계:1/3")
        progress_queue.put("파일:CINEMATIC_DIALOGUE.xlsm")

        # 데이터 읽기
        cinematic_data = read_excel_file(cinematic_path, sheet_name=1, header_row=1, skip_rows=9)
        # 글로벌 OnOFF=1 필터링 (G열, 인덱스 6)
        cinematic_data = cinematic_data[cinematic_data.iloc[:, 6] == 1]
        
        # 시간 계산 및 전송
        elapsed = int(time.time() - start_time)
        remaining = int((elapsed / 20) * 80) if elapsed > 0 else 0
        progress_queue.put(("time", elapsed, remaining))
        
        progress_queue.put(20)
        progress_queue.put("처리된 파일:1")

        progress_queue.put("파일:SMALLTALK_DIALOGUE.xlsm")
        smalltalk_data = read_excel_file(smalltalk_path, sheet_name=1, header_row=1, skip_rows=4)
        # 글로벌 OnOFF=1 필터링 (G열, 인덱스 6)
        smalltalk_data = smalltalk_data[smalltalk_data.iloc[:, 6] == 1]
        
        # 시간 계산 및 전송
        elapsed = int(time.time() - start_time)
        remaining = int((elapsed / 40) * 60) if elapsed > 0 else 0
        progress_queue.put(("time", elapsed, remaining))
        
        progress_queue.put(40)
        progress_queue.put("처리된 파일:2")

        # 단계 업데이트
        progress_queue.put("단계:2/3")
        progress_queue.put("파일:데이터 병합 중...")

        # 열 인덱스가 범위를 벗어나지 않는지 확인
        cinematic_cols = cinematic_data.shape[1]
        smalltalk_cols = smalltalk_data.shape[1]

        # 결과물 파일의 헤더 설정
        headers = ['#', 'Table Name', 'String ID', 'Table/ID', 'NPC ID', 'Speaker Name',
                   'KO (M)', 'KO (F)', 'EN (M)', 'EN (F)', 'CT (M)', 'CT (F)', 'CS (M)',
                   'CS (F)', 'JA (M)', 'JA (F)', 'TH (M)', 'TH (F)', 'ES-LATAM (M)', 'ES-LATAM (F)',
                   'PT-BR (M)', 'PT-BR (F)', 'NOTE']

        # 결과 데이터 프레임 생성 - 미리 열을 생성해둠
        result_df = pd.DataFrame(columns=headers)

        # 각 데이터프레임의 길이 확인
        cin_len = len(cinematic_data)
        small_len = len(smalltalk_data)
        total_len = cin_len + small_len

        # 결과 데이터프레임에 필요한 개수만큼 행 추가 (빈 행으로)
        result_df = pd.DataFrame(index=range(total_len), columns=headers)

        # 인덱스 열 채우기
        result_df['#'] = range(1, total_len + 1)

        # Table Name 열 채우기
        result_df.loc[:cin_len-1, 'Table Name'] = 'CINEMATIC_DIALOGUE'
        result_df.loc[cin_len:, 'Table Name'] = 'SMALLTALK_DIALOGUE'

        # 안전하게 열 인덱스 확인하고 데이터 할당
        # String ID 열 (인덱스 7)
        if 7 < cinematic_cols and 7 < smalltalk_cols:
            result_df.loc[:cin_len-1, 'String ID'] = cinematic_data.iloc[:, 7].values
            result_df.loc[cin_len:, 'String ID'] = smalltalk_data.iloc[:, 7].values

        # Table/ID 열 생성
        result_df['Table/ID'] = result_df['Table Name'] + '/' + result_df['String ID'].astype(str)

        # NPC ID 열 (인덱스 8)
        if 8 < cinematic_cols and 8 < smalltalk_cols:
            result_df.loc[:cin_len-1, 'NPC ID'] = cinematic_data.iloc[:, 8].values
            result_df.loc[cin_len:, 'NPC ID'] = smalltalk_data.iloc[:, 8].values

        # 나머지 언어 데이터 열 매핑
        language_mapping = {
            'KO (M)': (11, 12),
            'KO (F)': (12, 13),
            'EN (M)': (13, 14),
            'EN (F)': (14, 15),
            'CT (M)': (15, 16),
            'CT (F)': (16, 17),
            'CS (M)': (17, 18),
            'CS (F)': (18, 19),
            'JA (M)': (19, 20),
            'JA (F)': (20, 21),
            'TH (M)': (21, 22),
            'TH (F)': (22, 23),
            'ES-LATAM (M)': (23, 24),
            'ES-LATAM (F)': (24, 25),
            'PT-BR (M)': (25, 26),
            'PT-BR (F)': (26, 27),
            'NOTE': (29, 30)
        }

        # 각 언어 열에 데이터 안전하게 채우기
        for col_name, (cin_idx, small_idx) in language_mapping.items():
            # 인덱스가 범위 내에 있는지 확인
            if cin_idx < cinematic_cols and small_idx < smalltalk_cols:
                result_df.loc[:cin_len-1, col_name] = cinematic_data.iloc[:, cin_idx].values
                result_df.loc[cin_len:, col_name] = smalltalk_data.iloc[:, small_idx].values
            else:
                # 인덱스가 범위를 벗어나면 빈 값으로 설정
                result_df[col_name] = ''

        progress_queue.put(60)
        progress_queue.put("단계:3/3")
        progress_queue.put("파일:NPC.xlsm")

        # 원본 3(NPC.xlsm) 데이터 읽기 (두 번째 시트)
        npc_data = pd.read_excel(npc_path, sheet_name='NPC', header=1)
        npc_data = npc_data.drop_duplicates(subset=npc_data.columns[7])
        progress_queue.put("처리된 파일:3")

        # 결과 파일의 'NPC ID' 열과 원본 3의 'H열 유니크 아이디' 열을 기준으로 'J열 NPC 이름' 값을 불러오기
        # 안전하게 매핑을 위해 try-except 구문 사용
        try:
            # Dictionary 매핑을 생성
            npc_map = dict(zip(npc_data.iloc[:, 7], npc_data.iloc[:, 9]))
            # 매핑된 값이 없으면 원래 값을 유지
            result_df['Speaker Name'] = result_df['NPC ID'].map(npc_map).fillna(result_df['NPC ID'])
        except Exception as e:
            # 매핑 실패시 오류 메시지 표시
            progress_queue.put(("error", f"NPC 이름 매핑 중 오류 발생: {str(e)}"))
            return

        # 9번째 열이 빈 셀(NaN) 또는 0 또는 '미사용'인 행 제거
        # 'EN (M)' 열이 9번째 열에 해당
        result_df = result_df[~(pd.isna(result_df['EN (M)']) | result_df['EN (M)'].isin([0, '미사용']))]

        # 인덱스 열 갱신
        result_df['#'] = range(1, len(result_df) + 1)

        # 시간 계산 및 전송
        elapsed = int(time.time() - start_time)
        remaining = int((elapsed / 80) * 20) if elapsed > 0 else 0
        progress_queue.put(("time", elapsed, remaining))
        
        progress_queue.put(80)
        progress_queue.put("파일:결과 파일 저장 중...")

        # String ID를 정수로 변환 (소수점 제거)
        result_df['String ID'] = pd.to_numeric(result_df['String ID'], errors='coerce').fillna(0).astype('int64')
        # Table/ID 재생성 (정수 기반)
        result_df['Table/ID'] = result_df['Table Name'] + '/' + result_df['String ID'].astype(str)

        # 출력 파일 이름 설정
        date_str = datetime.datetime.now().strftime('%m%d')
        output_file = f'{date_str}_MIR4_MASTER_DIALOGUE.xlsx'

        # 파일이 이미 존재할 경우 다른 이름으로 저장
        counter = 1
        while os.path.exists(output_file):
            output_file = f'{date_str}_MIR4_MASTER_DIALOGUE_{counter}.xlsx'
            counter += 1

        # 결과 파일 저장 (엑셀)
        result_df.to_excel(output_file, index=False)

        # 결과 파일 서식 지정
        wb = load_workbook(output_file)
        ws = wb.active

        # 폰트 및 서식 설정
        header_font = Font(name='맑은 고딕', size=12, bold=True, color='9C5700')
        default_font = Font(name='맑은 고딕', size=10)
        header_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        border_style = Side(border_style='thin', color='000000')
        full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        # 헤더 행 서식 지정
        if isinstance(ws, Worksheet):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = full_border

            # 나머지 셀 서식 지정
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = default_font
                    cell.border = full_border

            # 틀 고정
            ws.freeze_panes = 'A2'

        # 서식 지정된 파일 저장
        wb.save(output_file)

        # 결과 파일 읽기 전용 설정
        os.chmod(output_file, stat.S_IREAD)

        progress_queue.put(100)

        elapsed_time = time.time() - start_time
        progress_queue.put(f"완료:파일이 {output_file}로 저장되었습니다. 소요 시간: {int(elapsed_time)}초")

    except Exception as e:
        progress_queue.put(("error", str(e)))
