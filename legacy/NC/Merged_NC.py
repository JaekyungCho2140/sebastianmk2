import sys
import platform

# ProcessPoolExecutor에서 사용할 함수는 반드시 글로벌로 정의해야 함

def read_excel_file(file_path):
    import pandas as pd
    return pd.read_excel(file_path)

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog
except ImportError as e:
    print("[오류] tkinter가 설치되어 있지 않습니다.")
    os_name = platform.system()
    if os_name == "Linux":
        print("리눅스 환경에서는 아래 명령어로 설치하세요:")
        print("  Ubuntu/Debian: sudo apt-get install python3-tk")
        print("  RedHat/CentOS/Fedora: sudo dnf install python3-tkinter")
    elif os_name == "Windows":
        print("Windows에서는 Python 설치 시 'tcl/tk and IDLE' 옵션이 선택되어야 합니다. Python을 재설치해보세요.")
    elif os_name == "Darwin":
        print("Mac에서는 Python 공식 설치본에 포함되어 있습니다. 문제가 계속되면 Python을 재설치해보세요.")
    else:
        print("운영체제에 맞는 tkinter 설치 방법을 확인해 주세요.")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("[오류] pandas가 설치되어 있지 않습니다. 'pip install pandas'로 설치해 주세요.")
    sys.exit(1)

import os
from datetime import datetime
import time

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("[오류] openpyxl이 설치되어 있지 않습니다. 'pip install openpyxl'로 설치해 주세요.")
    sys.exit(1)

import logging

try:
    from progress_window import ProgressWindow
except ImportError:
    print("[오류] progress_window.py 파일이 현재 폴더에 없거나, PYTHONPATH에 없습니다. 파일 위치와 이름을 확인해 주세요.")
    sys.exit(1)

try:
    from PIL import Image, ImageTk
except ImportError:
    print("[오류] Pillow(PIL) 라이브러리가 필요합니다. 'pip install pillow'로 설치해 주세요.")
    sys.exit(1)

from concurrent.futures import ProcessPoolExecutor

# 리소스 경로 함수 (PyInstaller 호환)
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# 로깅 설정
logging.basicConfig(filename='l10n_merger.log', level=logging.DEBUG, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

class L10nTableGrabber:
    def __init__(self, master):
        self.master = master
        master.title("NC Merged")
        master.geometry("480x640")
        master.resizable(False, False)

        # 배경 이미지 설정 (Pillow 사용)
        bg_img_path = resource_path("cho_NC_480x640.jpg")
        pil_bg_image = Image.open(bg_img_path)
        self.bg_image = ImageTk.PhotoImage(pil_bg_image)
        self.bg_label = tk.Label(master, image=self.bg_image)
        self.bg_label.place(x=0, y=0, relwidth=1, relheight=1)

        # 버튼 이미지 로드 (Pillow 사용)
        self.btn_start_img = ImageTk.PhotoImage(Image.open(resource_path("cho_NC_button_S_60x25.jpg")))
        self.btn_start_img_hover = ImageTk.PhotoImage(Image.open(resource_path("cho_NC_button_S_click_60x25.jpg")))
        self.btn_end_img = ImageTk.PhotoImage(Image.open(resource_path("cho_NC_button_E_50x25.jpg")))
        self.btn_end_img_hover = ImageTk.PhotoImage(Image.open(resource_path("cho_NC_button_E_click_50x25.jpg")))

        # 실행 버튼
        self.run_button = tk.Label(master, image=self.btn_start_img, bd=0, cursor="hand2")
        self.run_button.place(x=115, y=505, width=60, height=25)
        self.run_button.bind("<Button-1>", lambda e: self.run_program())
        self.run_button.bind("<Enter>", lambda e: self.run_button.config(image=self.btn_start_img_hover))
        self.run_button.bind("<Leave>", lambda e: self.run_button.config(image=self.btn_start_img))

        # 종료 버튼
        self.exit_button = tk.Label(master, image=self.btn_end_img, bd=0, cursor="hand2")
        self.exit_button.place(x=315, y=505, width=50, height=25)
        self.exit_button.bind("<Button-1>", lambda e: master.quit())
        self.exit_button.bind("<Enter>", lambda e: self.exit_button.config(image=self.btn_end_img_hover))
        self.exit_button.bind("<Leave>", lambda e: self.exit_button.config(image=self.btn_end_img))

        self.progress_window = None

    def run_program(self):
        folder_path = filedialog.askdirectory()
        if not self.check_files(folder_path):
            messagebox.showerror("오류", "올바른 폴더가 아닙니다.")
            return

        date = self.get_date_input()
        if not date:
            return

        milestone = self.get_milestone_input()
        if not milestone:
            return

        self.process_files(folder_path, date, milestone)

    def check_files(self, folder_path):
        required_files = [
            "StringEnglish.xlsx", "StringJapanese.xlsx", "StringPortuguese.xlsx",
            "StringRussian.xlsx", "StringSimplifiedChinese.xlsx", "StringSpanish.xlsx",
            "StringThai.xlsx", "StringTraditionalChinese.xlsx"
        ]
        return all(os.path.isfile(os.path.join(folder_path, file)) for file in required_files)

    def get_date_input(self):
        date = simpledialog.askstring("날짜 입력", "YYMMDD 형식으로 업데이트일을 입력하세요.")
        if not date or not date.isdigit() or len(date) != 6:
            messagebox.showerror("오류", "유효한 값이 아닙니다.")
            return None
        return date

    def get_milestone_input(self):
        milestone = simpledialog.askstring("마일스톤 입력", "마일스톤 차수를 입력하세요.")
        if not milestone or not milestone.isdigit() or len(milestone) > 3:
            messagebox.showerror("오류", "유효한 값이 아닙니다.")
            return None
        return milestone

    def process_files(self, folder_path, date, milestone):
        start_time = time.time()
        
        file_names = [
            "StringEnglish.xlsx", "StringTraditionalChinese.xlsx", "StringSimplifiedChinese.xlsx",
            "StringJapanese.xlsx", "StringThai.xlsx", "StringSpanish.xlsx",
            "StringPortuguese.xlsx", "StringRussian.xlsx"
        ]
        
        dfs = []
        total_steps = len(file_names) + 3  # 파일 읽기 + 병합 + 저장 + 서식 지정
        current_step = 0

        # ProgressWindow 초기화
        self.progress_window = ProgressWindow(
            self.master,
            title="파일 처리 중",
            theme_color="#4CAF50"
        )

        def process_worker(queue):
            nonlocal current_step
            import concurrent.futures

            t0 = time.time()
            with ProcessPoolExecutor() as executor:
                file_paths = [os.path.join(folder_path, f) for f in file_names]
                results = list(executor.map(read_excel_file, file_paths))
                dfs.extend(results)
                for idx, file_name in enumerate(file_names):
                    queue.put(f"파일:{file_name}")
                    queue.put(f"단계:{idx + 1}/{total_steps}")
                    queue.put(f"처리된 파일:{idx + 1}")
                    current_step += 1
                    queue.put(int(current_step / total_steps * 100))
            t1 = time.time()
            print(f"전체 파일 병렬 읽기 소요 시간: {t1 - t0:.2f}초")

            # 첫 번째 파일에서 기본 열 (Key, Source 등)을 가져옴
            result_df = dfs[0][['Key', 'Source', 'Comment', 'TableName', 'Status']]

            # 각 언어별 Target 열 추가 (concat으로 병합)
            lang_codes = ['EN', 'CT', 'CS', 'JA', 'TH', 'ES', 'PT', 'RU']
            target_dfs = [dfs[i][['Target']].rename(columns={'Target': f'Target_{lang_codes[i]}'}) for i in range(len(dfs))]
            result_df = pd.concat([result_df] + target_dfs, axis=1)

            # 메모리 사용량 최적화: 불필요한 데이터가 남아있지 않은지 확인
            # (필요 없는 컬럼/행이 있다면 이 시점에서 제거)

            # Python의 None 값을 빈 문자열로 대체 (텍스트로 'None'은 그대로 유지)
            for col in result_df.columns:
                if col != 'Comment':
                    result_df[col] = result_df[col].apply(lambda x: 'None' if pd.isna(x) else x)

            # NaN, inf, -inf를 모두 빈 문자열로 변환 (xlsxwriter 오류 방지)
            import numpy as np
            result_df = result_df.replace([np.nan, np.inf, -np.inf], '', regex=False)

            # 열 순서 재정렬
            result_df = result_df[['Key', 'Source', 'Target_EN', 'Target_CT', 'Target_CS', 
                                'Target_JA', 'Target_TH', 'Target_ES', 'Target_PT', 
                                'Target_RU', 'Comment', 'TableName', 'Status']]

            logging.info(f"Result DataFrame shape: {result_df.shape}")
            print(f"Result DataFrame shape: {result_df.shape}")

            current_step += 1
            queue.put(int(current_step / total_steps * 100))

            # 저장 및 서식 적용 (xlsxwriter)
            t_save_start = time.time()
            output_file = f"{date}_M{milestone}_StringALL.xlsx"
            output_path = os.path.join(folder_path, output_file)
            try:
                import xlsxwriter
                workbook = xlsxwriter.Workbook(output_path)
                worksheet = workbook.add_worksheet('Sheet1')

                # 헤더 스타일 (가운데 정렬)
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'vcenter',
                    'align': 'center',
                    'fg_color': '#DAE9F8',
                    'font_name': '맑은 고딕',
                    'font_size': 10,
                    'border': 1
                })
                # 데이터 셀 스타일 (왼쪽 정렬 + 텍스트 서식)
                cell_format = workbook.add_format({
                    'font_name': '맑은 고딕',
                    'font_size': 10,
                    'align': 'left',
                    'valign': 'vcenter',
                    'num_format': '@'  # 텍스트 서식
                })
                for col_num, value in enumerate(result_df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                    worksheet.set_column(col_num, col_num, 24, cell_format)

                for row_num in range(len(result_df)):
                    for col_num in range(len(result_df.columns)):
                        value = result_df.iloc[row_num, col_num]
                        worksheet.write_string(row_num + 1, col_num, str(value), cell_format)

                workbook.close()
                logging.info(f"Successfully saved result to {output_path}")
                print(f"Successfully saved result to {output_path}")
            except Exception as e:
                logging.error(f"Error saving result: {str(e)}")
                queue.put(("error", f"결과 저장 실패: {str(e)}"))
                return
            t_save_end = time.time()
            print(f"엑셀 저장 및 서식 적용 소요 시간: {t_save_end - t_save_start:.2f}초")

            current_step += 1
            queue.put(100)
            queue.put("완료:테이블 병합을 완료했습니다.")

        # ProgressWindow 시작
        self.progress_window.start(
            process_worker,
            total_steps=total_steps,
            total_files=len(file_names)
        )

if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    root = tk.Tk()
    app = L10nTableGrabber(root)
    root.mainloop()

    print(sys.executable)
    import tkinter
    print(tkinter.TkVersion)