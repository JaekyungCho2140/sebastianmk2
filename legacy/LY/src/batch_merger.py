"""
배치별 병합 모듈

PRD v1.3.0 섹션 2 "Merge Batches"에 정의된 알고리즘을 구현합니다.
여러 배치의 언어별 파일을 병합하고 중복 KEY를 제거합니다.
"""

import re
from typing import Dict, List, Tuple, Optional
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook

from .validator import LANGUAGE_ORDER

VALID_LANGUAGES = ['EN', 'CT', 'CS', 'JA', 'TH', 'PT-BR', 'RU']
from .error_messages import get_user_friendly_message, format_batch_duplicates
from .excel_format import apply_split_format


# 배치 폴더명 패턴 (PRD 섹션 2.2.1)
BATCH_FOLDER_PATTERN = re.compile(r'^(\d{6})_(REGULAR|EXTRA(\d{1,2}))$')

# 배치 파일명 패턴 (PRD 섹션 2.2.2)
BATCH_FILE_PATTERN = re.compile(r'^(\d{6})_(EN|CT|CS|JA|TH|PT-BR|RU)_(.+)\.xlsx$')


class BatchMergerError(Exception):
    """배치 병합 오류"""
    def __init__(self, message: str, error_code: str = None):
        super().__init__(message)
        self.error_code = error_code


class UserCancelledError(Exception):
    """사용자 취소"""
    pass


def validate_batch_folder_name(folder_name: str) -> Tuple[bool, str, str]:
    """
    배치 폴더명 검증

    Args:
        folder_name: 폴더명

    Returns:
        (유효성, 날짜, 배치명)

    Reference:
        PRD 섹션 2.2.1
    """
    match = BATCH_FOLDER_PATTERN.match(folder_name)
    if not match:
        return False, '', ''

    date = match.group(1)
    batch_name = match.group(2)

    # EXTRA 번호 범위 확인 (0~20)
    if batch_name.startswith('EXTRA'):
        extra_num = int(match.group(3))
        if extra_num < 0 or extra_num > 20:
            return False, '', ''

    return True, date, batch_name


def extract_language_from_filename(filename: str, expected_batch: str) -> Optional[str]:
    """
    파일명에서 언어 코드 추출

    Args:
        filename: 파일명
        expected_batch: 예상 배치명 (폴더명 기준)

    Returns:
        언어 코드 또는 None

    Reference:
        PRD 섹션 2.2.2
    """
    match = BATCH_FILE_PATTERN.match(filename)
    if not match:
        return None

    date = match.group(1)
    lang_code = match.group(2)
    batch_name = match.group(3)

    # 배치명 일치 확인
    if batch_name != expected_batch:
        return None

    # 언어 코드 유효성 확인
    if lang_code not in VALID_LANGUAGES:
        return None

    return lang_code


def sort_batches(batch_names: List[str]) -> List[str]:
    """
    배치명 정렬

    REGULAR가 항상 첫 번째, EXTRA는 번호 오름차순

    Args:
        batch_names: 배치명 리스트

    Returns:
        정렬된 배치명 리스트

    Reference:
        PRD 섹션 2.3.1
    """
    regular = []
    extras = []

    for batch in batch_names:
        if batch == 'REGULAR':
            regular.append(batch)
        elif batch.startswith('EXTRA'):
            # EXTRA 번호 추출
            num = int(batch[5:])  # 'EXTRA' 이후 숫자
            extras.append((num, batch))

    # EXTRA 번호순 정렬
    extras.sort(key=lambda x: x[0])

    # REGULAR + EXTRA 순서
    return regular + [b for _, b in extras]


def apply_status_completion(final_data: Dict[str, List[List]]) -> Dict[str, List[List]]:
    """
    Status 자동 완료 처리

    '번역필요', '수정' 상태를 '완료'로 변경합니다.

    Args:
        final_data: {언어코드: [[헤더], [행1], [행2], ...]}

    Returns:
        Status 변경된 데이터

    Reference:
        PRD v1.4.0 섹션 1.2
    """
    STATUS_MAPPING = {
        "번역필요": "완료",
        "수정": "완료"
    }

    for lang_code in final_data.keys():
        rows = final_data[lang_code]

        # 헤더 제외하고 처리
        for row_idx in range(1, len(rows)):
            row = rows[row_idx]

            if len(row) >= 5:  # Status는 5번째 컬럼 (E열, 인덱스 4)
                current_status = row[4]

                # 매핑에 있으면 변경
                if current_status in STATUS_MAPPING:
                    row[4] = STATUS_MAPPING[current_status]

    return final_data


def sort_batches_with_base(batch_names: List[str], base_batch: str) -> List[str]:
    """
    배치명 정렬 (기준 배치 우선)

    Args:
        batch_names: 배치명 리스트
        base_batch: 기준 배치명

    Returns:
        정렬된 배치명 리스트 (base_batch가 첫 번째)

    Reference:
        PRD v1.4.0 섹션 2.5
    """
    # 기준 배치 제외
    other_batches = [b for b in batch_names if b != base_batch]

    # REGULAR/EXTRA 분류
    regular = []
    extras = []

    for batch in other_batches:
        if batch == 'REGULAR':
            regular.append(batch)
        elif batch.startswith('EXTRA'):
            num = int(batch[5:])
            extras.append((num, batch))

    # EXTRA 번호순 정렬
    extras.sort(key=lambda x: x[0])

    # 기준 배치 + REGULAR (있으면) + EXTRA 순
    sorted_others = regular + [b for _, b in extras]

    return [base_batch] + sorted_others


def scan_batch_folders(root_folder: Path) -> Dict:
    """
    배치 폴더 스캔 및 검증

    Args:
        root_folder: 루트 폴더 경로

    Returns:
        {
            'REGULAR': {
                'folder': '251126_REGULAR',
                'date': '251126',
                'files': {'EN': '251126_EN_REGULAR.xlsx', ...},
                'valid': True,
                'file_count': 7,
                'missing_languages': []
            },
            ...
        }

    Raises:
        BatchMergerError: 스캔 실패 시

    Reference:
        PRD 섹션 2.4.2
    """
    if not root_folder.exists():
        raise BatchMergerError(
            get_user_friendly_message("FILE_NOT_FOUND", path=str(root_folder)),
            "FILE_NOT_FOUND"
        )

    batch_info = {}

    # 하위 폴더 목록 조회 (직접 하위만, 재귀 없음)
    subdirs = [d for d in root_folder.iterdir() if d.is_dir()]

    for folder in subdirs:
        # 폴더명 검증
        is_valid, date, batch_name = validate_batch_folder_name(folder.name)

        if not is_valid:
            # 패턴 불일치 폴더는 무시 (Output 등)
            continue

        # 배치 내 파일 검증
        file_info = validate_batch_files(folder, batch_name)

        batch_info[batch_name] = {
            'folder': folder.name,
            'date': date,
            'files': file_info['files'],
            'valid': file_info['valid'],
            'file_count': file_info['file_count'],
            'missing_languages': file_info['missing_languages'],
            'duplicate_languages': file_info.get('duplicate_languages', [])
        }

    # REGULAR 폴더 존재 확인
    if 'REGULAR' not in batch_info:
        raise BatchMergerError(
            get_user_friendly_message("BATCH_REGULAR_MISSING"),
            "BATCH_REGULAR_MISSING"
        )

    return batch_info


def validate_batch_files(batch_folder: Path, batch_name: str) -> Dict:
    """
    배치 폴더 내 파일 검증

    Args:
        batch_folder: 배치 폴더 경로
        batch_name: 배치명

    Returns:
        {
            'files': {'EN': '파일명.xlsx', ...},
            'valid': True/False,
            'file_count': 숫자,
            'missing_languages': [...],
            'duplicate_languages': [...]
        }

    Reference:
        PRD 섹션 2.2.3
    """
    # 직접 하위 파일만 스캔 (재귀 없음)
    files = [f for f in batch_folder.iterdir() if f.is_file() and f.suffix == '.xlsx']

    lang_files = {}
    lang_counts = {lang: 0 for lang in VALID_LANGUAGES}

    for file in files:
        lang = extract_language_from_filename(file.name, batch_name)
        if lang:
            lang_counts[lang] += 1
            if lang_counts[lang] == 1:
                lang_files[lang] = file.name

    # 중복 언어 확인
    duplicates = [lang for lang, count in lang_counts.items() if count > 1]

    # 누락 언어 확인
    missing = [lang for lang in VALID_LANGUAGES if lang_counts[lang] == 0]

    return {
        'files': lang_files,
        'valid': len(lang_files) == 7 and len(duplicates) == 0,
        'file_count': len(lang_files),
        'missing_languages': missing,
        'duplicate_languages': duplicates
    }


def validate_batch_selection(
    selected_batches: List[str],
    base_batch: Optional[str],
    batch_info: Dict
) -> Tuple[bool, str]:
    """
    배치 선택 검증

    Args:
        selected_batches: 선택된 배치명 리스트
        base_batch: 기준 배치명 (None이면 미선택)
        batch_info: scan_batch_folders() 결과

    Returns:
        (유효성, 오류 메시지)

    Reference:
        PRD v1.4.0 섹션 2.4
    """
    # 1. 기준 배치 선택 확인
    if not base_batch:
        return False, (
            "기준 배치를 선택해주세요.\n\n"
            "기준 배치는 적재 순서의 첫 번째가 되는 배치입니다.\n"
            "라디오 버튼으로 선택해주세요."
        )

    # 2. 기준 배치가 선택된 배치에 포함되는지 확인
    if base_batch not in selected_batches:
        return False, f"기준 배치 {base_batch}가 선택되지 않았습니다."

    # 3. 최소 선택 수 (2개 이상)
    if len(selected_batches) < 2:
        return False, get_user_friendly_message("SELECTION_TOO_FEW", count=len(selected_batches))

    # 4. 기준 배치 완전성 검증
    if batch_info[base_batch]['file_count'] != 7:
        missing = batch_info[base_batch]['missing_languages']
        return False, get_user_friendly_message(
            "BATCH_INCOMPLETE",
            batch=base_batch,
            missing=', '.join(missing)
        )

    # 5. 선택된 배치 완전성 검증
    for batch in selected_batches:
        if not batch_info[batch]['valid']:
            # 중복 파일 확인
            if batch_info[batch]['duplicate_languages']:
                dup_lang = batch_info[batch]['duplicate_languages'][0]
                return False, get_user_friendly_message(
                    "BATCH_FILE_DUPLICATE",
                    batch=batch,
                    lang=dup_lang,
                    files="(파일 목록)"
                )

            # 누락 파일
            missing = batch_info[batch]['missing_languages']
            return False, get_user_friendly_message(
                "BATCH_INCOMPLETE",
                batch=batch,
                missing=', '.join(missing)
            )

    return True, ''


def merge_batches_for_language(
    language_code: str,
    selected_batches: List[str],
    batch_info: Dict,
    root_folder: Path,
    cancel_check=None
) -> List[List]:
    """
    특정 언어의 모든 배치 데이터를 순차 적재

    Args:
        language_code: 언어 코드 (EN, CT, ...)
        selected_batches: 선택된 배치 목록 (정렬됨)
        batch_info: 배치 정보
        root_folder: 루트 폴더 경로
        cancel_check: 취소 확인 함수 (optional)

    Returns:
        병합된 행 리스트 [[헤더], [행1], [행2], ...]

    Raises:
        BatchMergerError: 파일 읽기 실패 등
        UserCancelledError: 사용자 취소

    Reference:
        PRD 섹션 2.4.4
    """
    merged_rows = []

    for batch_idx, batch_name in enumerate(selected_batches):
        # 취소 확인
        if cancel_check and cancel_check():
            raise UserCancelledError("사용자가 작업을 취소했습니다.")

        batch = batch_info[batch_name]
        file_path = root_folder / batch['folder'] / batch['files'][language_code]

        # 파일 로드
        try:
            wb = load_workbook(file_path)
        except Exception as e:
            raise BatchMergerError(
                get_user_friendly_message("FILE_READ_ERROR", file=file_path.name, error=str(e)),
                "FILE_READ_ERROR"
            )

        ws = wb.active

        # 헤더 검증 (첫 배치만)
        if batch_idx == 0:
            expected_headers = ['Table', 'KEY', 'Source', 'Target', 'Status', 'NOTE', 'Date']
            actual_headers = [cell.value for cell in ws[1]]

            if actual_headers != expected_headers:
                raise BatchMergerError(
                    get_user_friendly_message(
                        "INVALID_HEADERS",
                        file=file_path.name,
                        expected=expected_headers,
                        actual=actual_headers
                    ),
                    "INVALID_HEADERS"
                )

            # 첫 배치: 헤더 추가
            merged_rows.append(list(actual_headers))

            # 데이터 추가 (빈 행 스킵)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7 and row[1] and str(row[1]).strip():  # KEY 있는 행만
                    merged_rows.append(list(row[:7]))
        else:
            # 이후 배치: 헤더 제외, 데이터만 추가 (2행부터)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7:  # 최소 7개 컬럼 있어야 함
                    # 빈 행 스킵 (KEY가 없으면 빈 행)
                    if row[1] and str(row[1]).strip():  # B열 KEY 확인
                        merged_rows.append(list(row[:7]))  # 7개 컬럼만 추출

    return merged_rows


def find_duplicates_within_batch(
    data_rows: List[List],
    selected_batches: List[str],
    batch_row_counts: Dict[str, int]
) -> Dict[str, List[str]]:
    """
    배치 내 중복 KEY 검출

    Args:
        data_rows: 전체 데이터 행 (헤더 제외)
        selected_batches: 선택된 배치 목록 (정렬됨)
        batch_row_counts: {배치명: 행 수}

    Returns:
        {배치명: [중복KEY1, 중복KEY2, ...]}

    Reference:
        PRD 섹션 2.4.5
    """
    batch_duplicates = {}
    current_idx = 0

    for batch_name in selected_batches:
        row_count = batch_row_counts[batch_name]
        batch_rows = data_rows[current_idx:current_idx + row_count]

        # KEY 중복 검출
        key_counts = {}
        for row in batch_rows:
            if len(row) >= 2:
                key = row[1]  # B열 (KEY)
                if key:
                    key_counts[key] = key_counts.get(key, 0) + 1

        # 중복 KEY 추출
        duplicates = [k for k, count in key_counts.items() if count > 1]

        if duplicates:
            batch_duplicates[batch_name] = duplicates

        current_idx += row_count

    return batch_duplicates


def parse_and_validate_date(date_str: str, key: str, row_idx: int) -> datetime:
    """
    Date 문자열 파싱 및 검증

    Args:
        date_str: Date 문자열
        key: KEY 값
        row_idx: 행 번호

    Returns:
        파싱된 datetime 객체

    Raises:
        BatchMergerError: Date 검증 실패 시

    Reference:
        PRD 섹션 2.5
    """
    # 1. 빈 Date 검증
    if not date_str or str(date_str).strip() == '':
        raise BatchMergerError(
            get_user_friendly_message("DATE_EMPTY_IN_DUPLICATE", key=key, row=row_idx),
            "DATE_EMPTY_IN_DUPLICATE"
        )

    # 2. Date 형식 파싱 (YYYY-MM-DD HH:MM)
    try:
        parsed_date = datetime.strptime(str(date_str).strip(), "%Y-%m-%d %H:%M")
    except ValueError:
        raise BatchMergerError(
            get_user_friendly_message("DATE_FORMAT_INVALID", key=key, row=row_idx, date=date_str),
            "DATE_FORMAT_INVALID"
        )

    return parsed_date


def select_latest_row(key: str, rows: List[Tuple[int, List]]) -> Tuple[int, List]:
    """
    중복 KEY 중 Date 기준 최신 행 선택

    Args:
        key: KEY 값
        rows: [(원본행번호, 행데이터), ...]

    Returns:
        (원본행번호, 최신행데이터)

    Raises:
        BatchMergerError: Date 검증 실패 시

    Reference:
        PRD 섹션 2.4.5
    """
    # Date 파싱 및 검증
    parsed_rows = []

    for row_idx, row_data in rows:
        date_str = row_data[6]  # G열 (Date)
        parsed_date = parse_and_validate_date(date_str, key, row_idx)
        parsed_rows.append((row_idx, row_data, parsed_date))

    # Date 동일 검증
    dates = [d for _, _, d in parsed_rows]
    if len(set(dates)) != len(dates):
        # Date가 동일한 행이 있음
        date_counts = {}
        for d in dates:
            date_counts[d] = date_counts.get(d, 0) + 1

        duplicate_dates = [d.strftime("%Y-%m-%d %H:%M") for d, count in date_counts.items() if count > 1]

        raise BatchMergerError(
            get_user_friendly_message("DUPLICATE_DATE_SAME", key=key, date=', '.join(duplicate_dates)),
            "DUPLICATE_DATE_SAME"
        )

    # 최신 행 선택 (Date가 가장 큰 것)
    latest = max(parsed_rows, key=lambda x: x[2])

    return (latest[0], latest[1])


def remove_duplicate_keys(language_data: Dict[str, List[List]], selected_batches: List[str], batch_row_counts: Dict[str, int]) -> Tuple[Dict[str, List[List]], List[Dict]]:
    """
    중복 KEY 제거 (EN 기준 통합 검증)

    Args:
        language_data: {언어코드: [[헤더], [행1], ...]}
        selected_batches: 선택된 배치 목록
        batch_row_counts: {배치명: 행 수}

    Returns:
        (중복 제거된 데이터, 중복 제거 로그)

    Raises:
        BatchMergerError: 중복 검증 실패 시

    Reference:
        PRD 섹션 2.4.5
    """
    # 1. EN 파일에서 처리
    en_rows = language_data['EN']
    en_header = en_rows[0]
    en_data_rows = en_rows[1:]  # 헤더 제외

    # 2. 배치 내 중복 검출 (오류)
    batch_duplicates = find_duplicates_within_batch(en_data_rows, selected_batches, batch_row_counts)
    if batch_duplicates:
        raise BatchMergerError(
            get_user_friendly_message(
                "BATCH_INTERNAL_DUPLICATE",
                batch=list(batch_duplicates.keys())[0],
                keys='\n'.join(f"- {k}" for k in batch_duplicates[list(batch_duplicates.keys())[0]][:10])
            ),
            "BATCH_INTERNAL_DUPLICATE"
        )

    # 3. 순차 스캔하여 중복 검출 및 제거 (순차 적재 순서 유지)
    seen_keys = {}  # {KEY: (행_인덱스, 행_데이터, Date)}
    rows_to_remove = set()  # 제거할 행 인덱스
    duplicate_log = []

    for row_idx, row in enumerate(en_data_rows, start=2):
        if len(row) < 2:
            continue

        key = row[1]  # B열 (KEY)
        if not key:
            continue

        if key not in seen_keys:
            # 첫 등장: 기록
            seen_keys[key] = (row_idx, row, row[6])
        else:
            # 중복 발견: Date 비교
            prev_row_idx, prev_row, prev_date = seen_keys[key]
            curr_date = row[6]

            # Date 파싱 및 비교
            prev_parsed = parse_and_validate_date(prev_date, key, prev_row_idx)
            curr_parsed = parse_and_validate_date(curr_date, key, row_idx)

            if prev_parsed == curr_parsed:
                # Date 동일: 오류
                raise BatchMergerError(
                    get_user_friendly_message("DUPLICATE_DATE_SAME", key=key, date=prev_date),
                    "DUPLICATE_DATE_SAME"
                )

            # 최신 행 결정
            if curr_parsed > prev_parsed:
                # 현재 행이 더 최신: 이전 행 제거, 현재 행 유지
                rows_to_remove.add(prev_row_idx)

                # 로그 갱신
                existing_log = next((l for l in duplicate_log if l['key'] == key), None)
                if existing_log:
                    existing_log['removed'].append((prev_row_idx, prev_date))
                    existing_log['kept_row'] = row_idx
                    existing_log['kept_date'] = curr_date
                else:
                    duplicate_log.append({
                        'key': key,
                        'kept_row': row_idx,
                        'kept_date': curr_date,
                        'removed': [(prev_row_idx, prev_date)]
                    })

                seen_keys[key] = (row_idx, row, curr_date)
            else:
                # 이전 행이 더 최신: 현재 행 제거, 이전 행 유지
                rows_to_remove.add(row_idx)

                # 로그 갱신
                existing_log = next((l for l in duplicate_log if l['key'] == key), None)
                if existing_log:
                    existing_log['removed'].append((row_idx, curr_date))
                else:
                    duplicate_log.append({
                        'key': key,
                        'kept_row': prev_row_idx,
                        'kept_date': prev_date,
                        'removed': [(row_idx, curr_date)]
                    })

    # 4. 제거할 행 제외하고 최종 행 생성 (순차 적재 순서 그대로 유지)
    final_en_rows = [en_header]
    for row_idx, row in enumerate(en_data_rows, start=2):
        if row_idx not in rows_to_remove:
            final_en_rows.append(row)

    # 5. 다른 언어도 동일하게 중복 제거 (EN과 동일한 행만 제거)
    final_data = {'EN': final_en_rows}

    for lang in ['CT', 'CS', 'JA', 'TH', 'PT-BR', 'RU']:
        lang_header = language_data[lang][0]
        lang_data_rows = language_data[lang][1:]

        # EN에서 제거된 행 인덱스와 동일한 인덱스의 행 제거
        lang_final_rows = [lang_header]
        for row_idx, row in enumerate(lang_data_rows, start=2):
            if row_idx not in rows_to_remove:
                lang_final_rows.append(row)

        final_data[lang] = lang_final_rows

        # 언어별 행 수 검증
        lang_count = len(lang_final_rows) - 1
        en_count = len(final_en_rows) - 1
        if lang_count != en_count:
            raise BatchMergerError(
                get_user_friendly_message(
                    "LANGUAGE_ROW_COUNT_MISMATCH",
                    lang=lang,
                    en_count=en_count,
                    lang_count=lang_count
                ),
                "LANGUAGE_ROW_COUNT_MISMATCH"
            )

    return final_data, duplicate_log


def save_merged_batches(
    final_data: Dict[str, List[List]],
    output_folder: Path,
    date_prefix: str,
    overwrite_callback=None
) -> Dict[str, str]:
    """
    최종 병합 데이터 저장

    Args:
        final_data: 언어별 최종 데이터
        output_folder: 출력 폴더 경로
        date_prefix: 파일명 날짜 (YYMMDD)
        overwrite_callback: 덮어쓰기 확인 콜백

    Returns:
        {언어코드: 파일경로}

    Raises:
        BatchMergerError: 저장 실패 시
        UserCancelledError: 사용자 취소 시

    Reference:
        PRD 섹션 2.4.6
    """
    # 1. Output 폴더 생성
    try:
        output_folder.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise BatchMergerError(
            get_user_friendly_message("DIR_CREATE_ERROR", error=str(e)),
            "DIR_CREATE_ERROR"
        )

    # 2. 기존 파일 덮어쓰기 확인
    existing_files = []
    for lang in VALID_LANGUAGES:
        output_path = output_folder / f"{date_prefix}_{lang}.xlsx"
        if output_path.exists():
            existing_files.append(output_path.name)

    if existing_files and overwrite_callback:
        confirmed = overwrite_callback(existing_files)
        if not confirmed:
            raise UserCancelledError("사용자가 작업을 취소했습니다.")

    # 3. 파일 저장
    saved_files = {}

    for lang in VALID_LANGUAGES:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 데이터 작성
        for row in final_data[lang]:
            ws.append(row)

        # 서식 적용 (Split과 동일)
        apply_split_format(ws)

        # 저장
        output_path = output_folder / f"{date_prefix}_{lang}.xlsx"

        try:
            wb.save(output_path)
            saved_files[lang] = str(output_path)
        except Exception as e:
            # 저장 실패 시: 성공 파일 유지, 실패 메시지만
            saved_list = ', '.join(saved_files.keys()) if saved_files else '없음'
            raise BatchMergerError(
                f"{lang} 파일 저장에 실패했습니다.\n\n"
                f"경로: {output_path}\n"
                f"오류: {e}\n\n"
                f"디스크 공간이나 권한을 확인해주세요.\n"
                f"이미 저장된 파일: {saved_list}",
                "FILE_WRITE_ERROR"
            )

    return saved_files


def generate_merge_batches_log(log_info: Dict, output_folder: Path) -> Path:
    """
    로그 파일 생성

    Args:
        log_info: {
            'start_time': datetime,
            'end_time': datetime,
            'root_folder': Path,
            'selected_batches': list,
            'batch_processing': list,
            'duplicate_log': list,
            'final_stats': dict,
            'output_files': dict
        }

    Returns:
        로그 파일 경로

    Reference:
        PRD 섹션 2.4.7
    """
    timestamp = log_info['start_time'].strftime("%y%m%d_%H%M%S")
    log_path = output_folder / f"merge_batches_log_{timestamp}.txt"

    # 로그 내용 생성
    lines = []
    lines.append("=== Merge Batches 실행 로그 ===")
    lines.append(f"실행 시간: {log_info['start_time'].strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"루트 폴더: {log_info['root_folder']}")
    lines.append(f"선택 배치: {', '.join(log_info['selected_batches'])}")
    lines.append("")

    # 배치별 처리 내역
    for idx, batch_proc in enumerate(log_info['batch_processing'], start=1):
        lines.append(f"[{idx}/{len(log_info['selected_batches'])}] {batch_proc['batch']} 배치 처리 중...")
        for lang, row_count in batch_proc['languages'].items():
            lines.append(f"  - {lang}: {row_count:,}행 {'읽기' if idx == 1 else '적재'} 완료")
        lines.append("")

    # 중복 KEY 제거 내역
    duplicate_log = log_info['duplicate_log']
    if duplicate_log:
        lines.append("[중복 KEY 제거 내역]")
        lines.append(f"총 중복 KEY 발견: {len(duplicate_log)}개")
        lines.append("")

        for idx, dup in enumerate(duplicate_log, start=1):
            lines.append(f"  {idx}. KEY: {dup['key']}")
            for row_idx, removed_date in dup['removed']:
                lines.append(f"     - 제거: {removed_date} (행 {row_idx})")
            lines.append(f"     - 유지: {dup['kept_date']}")
            lines.append("")

    # 최종 결과
    stats = log_info['final_stats']
    lines.append("[최종 결과]")
    lines.append(f"  총 처리 행: {stats['total_rows']:,}행")
    lines.append(f"  중복 제거: {stats['duplicates_removed']}행")
    lines.append(f"  최종 행: {stats['final_rows']:,}행 (헤더 제외)")
    lines.append("")

    # 출력 파일
    lines.append("[출력 파일]")
    for lang, path in log_info['output_files'].items():
        row_count = len(log_info['final_data'][lang]) - 1  # 헤더 제외
        lines.append(f"  - {path} ({row_count:,}행)")
    lines.append("")

    # 완료 시간
    elapsed = (log_info['end_time'] - log_info['start_time']).total_seconds()
    minutes = int(elapsed // 60)
    seconds = int(elapsed % 60)
    time_str = f"{minutes}분 {seconds}초" if minutes > 0 else f"{seconds}초"

    lines.append(f"처리 완료: {log_info['end_time'].strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"소요 시간: {time_str}")
    lines.append("")
    lines.append("=" * 50)

    # 파일 저장
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    return log_path


def merge_batches(
    root_folder: Path,
    selected_batches: List[str],
    base_batch: str,
    batch_info: Dict,
    progress_callback=None,
    cancel_check=None,
    overwrite_callback=None
) -> Tuple[Dict[str, str], Path]:
    """
    배치 병합 메인 함수

    Args:
        root_folder: 루트 폴더 경로
        selected_batches: 선택된 배치 목록
        base_batch: 기준 배치명 (첫 번째로 적재됨)
        batch_info: 배치 정보
        progress_callback: 진행률 콜백 함수(percent, message)
        cancel_check: 취소 확인 함수 (returns bool)
        overwrite_callback: 덮어쓰기 확인 함수 (returns bool)

    Returns:
        (출력 파일 경로 딕셔너리, 로그 파일 경로)

    Raises:
        BatchMergerError: 처리 실패 시
        UserCancelledError: 사용자 취소 시

    Reference:
        PRD v1.4.0 섹션 2, 기능 1, 기능 2
    """
    start_time = datetime.now()

    # 배치 순서 정렬 (기준 배치 우선)
    sorted_batches = sort_batches_with_base(selected_batches, base_batch)

    # 로그 정보 초기화
    log_info = {
        'start_time': start_time,
        'root_folder': str(root_folder),
        'selected_batches': sorted_batches,
        'base_batch': base_batch,
        'batch_processing': [],
        'duplicate_log': [],
        'final_stats': {},
        'output_files': {},
        'final_data': {}
    }

    try:
        # Step 1: 배치 스캔 (이미 완료됨)
        if progress_callback:
            progress_callback(5, "배치 스캔 완료")

        # Step 3: 언어별 데이터 순차 적재
        language_data = {}
        batch_row_counts = {}

        total_batches = len(sorted_batches)
        load_weight = 40.0  # 40%

        for batch_idx, batch_name in enumerate(sorted_batches):
            batch_start_progress = 5 + (batch_idx / total_batches) * load_weight

            if progress_callback:
                progress_callback(
                    int(batch_start_progress),
                    f"{batch_name} 배치 읽기 중... ({batch_idx + 1}/{total_batches})"
                )

            # 배치 처리 로그
            batch_proc = {
                'batch': batch_name,
                'languages': {}
            }

            # 첫 배치 처리하여 행 수 파악
            first_lang_data = merge_batches_for_language(
                'EN', [batch_name], {batch_name: batch_info[batch_name]}, root_folder, cancel_check
            )

            if batch_idx == 0:
                # 첫 배치 (기준 배치): 데이터 초기화
                for lang in VALID_LANGUAGES:
                    language_data[lang] = merge_batches_for_language(
                        lang, [batch_name], {batch_name: batch_info[batch_name]}, root_folder, cancel_check
                    )
                    batch_proc['languages'][lang] = len(language_data[lang]) - 1  # 헤더 제외

                batch_row_counts[batch_name] = len(language_data['EN']) - 1
            else:
                # 이후 배치: 데이터 적재
                for lang in VALID_LANGUAGES:
                    new_data = merge_batches_for_language(
                        lang, [batch_name], {batch_name: batch_info[batch_name]}, root_folder, cancel_check
                    )
                    # 헤더 제외하고 기존 데이터에 추가
                    language_data[lang].extend(new_data[1:])
                    batch_proc['languages'][lang] = len(new_data) - 1

                batch_row_counts[batch_name] = len(new_data) - 1

            log_info['batch_processing'].append(batch_proc)

        if progress_callback:
            progress_callback(45, "모든 배치 읽기 완료")

        # Step 4: 중복 KEY 제거
        if progress_callback:
            progress_callback(50, "중복 KEY 제거 중...")

        final_data, duplicate_log = remove_duplicate_keys(language_data, sorted_batches, batch_row_counts)

        log_info['duplicate_log'] = duplicate_log
        log_info['final_data'] = final_data

        if progress_callback:
            progress_callback(75, "중복 제거 완료")

        # Step 4.5: Status 자동 완료 처리 (PRD v1.4.0 기능 1)
        if progress_callback:
            progress_callback(78, "Status 자동 완료 처리 중...")

        final_data = apply_status_completion(final_data)

        if progress_callback:
            progress_callback(80, "Status 처리 완료")

        # 최종 통계
        total_rows = sum(len(data) - 1 for data in language_data.values()) // 7  # 헤더 제외, 언어 평균
        final_rows = len(final_data['EN']) - 1
        duplicates_removed = len(duplicate_log)

        log_info['final_stats'] = {
            'total_rows': total_rows,
            'duplicates_removed': duplicates_removed,
            'final_rows': final_rows
        }

        # Step 5: 파일 저장
        if progress_callback:
            progress_callback(85, "파일 저장 중...")

        output_date = datetime.now().strftime("%y%m%d")
        output_dir = root_folder / "Output"

        saved_files = save_merged_batches(final_data, output_dir, output_date, overwrite_callback)

        log_info['output_files'] = saved_files

        if progress_callback:
            progress_callback(95, "로그 파일 생성 중...")

        # Step 6: 로그 파일 생성
        log_info['end_time'] = datetime.now()
        log_path = generate_merge_batches_log(log_info, output_dir)

        if progress_callback:
            progress_callback(100, "완료")

        return saved_files, log_path

    except (BatchMergerError, UserCancelledError):
        raise
    except Exception as e:
        raise BatchMergerError(
            f"예상치 못한 오류가 발생했습니다.\n\n"
            f"오류: {e}\n\n"
            f"로그를 확인하고 다시 시도해주세요.",
            "UNKNOWN_ERROR"
        )
