"""
Legacy Diff 모듈

PRD v1.4.0 섹션 3 "Legacy Diff"에 정의된 알고리즘을 구현합니다.
두 버전의 언어별 파일을 비교하여 Status가 "기존"인 행의 Target 변경사항을 추출합니다.
"""

from pathlib import Path
from typing import Dict, List, Tuple, Optional, Callable
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from .validator import ValidationError


# 지원 언어 목록
VALID_LANGUAGES = ['EN', 'CT', 'CS', 'JA', 'TH', 'PT-BR', 'RU']


class LegacyDiffError(Exception):
    """Legacy Diff 처리 오류"""
    def __init__(self, message: str, error_code: str = None):
        self.message = message
        self.error_code = error_code
        super().__init__(self.message)


def scan_language_files(folder: Path) -> Dict[str, Path]:
    """
    폴더에서 언어별 파일 스캔

    Args:
        folder: 스캔할 폴더

    Returns:
        {언어코드: Path 객체}

    파일명 패턴:
        - 유연한 매칭: 파일명에 언어 코드 포함 여부
        - 예: 251201_EN.xlsx, 251202_EN_REGULAR.xlsx 모두 EN으로 인식
    """
    files = {}
    xlsx_files = list(folder.glob("*.xlsx"))

    for lang in VALID_LANGUAGES:
        # 언어 코드가 파일명에 포함된 파일 찾기
        matching = [f for f in xlsx_files if f"_{lang}" in f.name or f"_{lang}." in f.name]

        if len(matching) == 1:
            files[lang] = matching[0]
        elif len(matching) > 1:
            # 중복 파일 - 가장 최근 수정된 파일 사용
            latest = max(matching, key=lambda f: f.stat().st_mtime)
            files[lang] = latest

    return files


def validate_diff_folders(folder1: Path, folder2: Path) -> Tuple[bool, str, Dict]:
    """
    Legacy Diff 폴더 검증

    Args:
        folder1: 비교1 폴더
        folder2: 비교2 폴더

    Returns:
        (유효성, 오류 메시지, 파일 매칭 정보)

    파일 매칭 정보:
        {
            'EN': (Path1, Path2),
            'CT': (Path1, Path2),
            ...
        }

    Reference:
        PRD v1.4.0 섹션 3.3
    """
    # 1. 비교1 폴더 파일 스캔
    files1 = scan_language_files(folder1)
    if len(files1) != 7:
        missing = set(VALID_LANGUAGES) - set(files1.keys())
        return False, (
            f"비교1 폴더에 7개 언어 파일이 필요합니다.\n\n"
            f"폴더: {folder1}\n"
            f"발견된 파일: {len(files1)}개\n"
            f"누락된 언어: {', '.join(missing) if missing else '없음'}\n\n"
            f"7개 언어 파일(EN, CT, CS, JA, TH, PT-BR, RU)을 확인해주세요."
        ), {}

    # 2. 비교2 폴더 파일 스캔
    files2 = scan_language_files(folder2)
    if len(files2) != 7:
        missing = set(VALID_LANGUAGES) - set(files2.keys())
        return False, (
            f"비교2 폴더에 7개 언어 파일이 필요합니다.\n\n"
            f"폴더: {folder2}\n"
            f"발견된 파일: {len(files2)}개\n"
            f"누락된 언어: {', '.join(missing) if missing else '없음'}\n\n"
            f"7개 언어 파일을 확인해주세요."
        ), {}

    # 3. 언어 매칭 확인
    lang_set1 = set(files1.keys())
    lang_set2 = set(files2.keys())

    if lang_set1 != lang_set2:
        missing_in_2 = lang_set1 - lang_set2
        missing_in_1 = lang_set2 - lang_set1

        msg = "두 폴더의 언어 파일이 일치하지 않습니다.\n\n"
        if missing_in_1:
            msg += f"비교1에 없음: {', '.join(missing_in_1)}\n"
        if missing_in_2:
            msg += f"비교2에 없음: {', '.join(missing_in_2)}\n"
        msg += "\n두 폴더 모두 동일한 언어 파일이 필요합니다."

        return False, msg, {}

    # 4. 파일 쌍 생성 (파일명 동일 여부는 강제하지 않음 - 다른 배치일 수 있음)
    file_pairs = {}
    for lang in VALID_LANGUAGES:
        file1 = files1[lang]
        file2 = files2[lang]
        file_pairs[lang] = (file1, file2)

    return True, "", file_pairs


def compare_language_files(
    file1: Path,
    file2: Path,
    language_code: str
) -> List[Dict]:
    """
    언어별 파일 비교

    Args:
        file1: 비교1 파일
        file2: 비교2 파일
        language_code: 언어 코드

    Returns:
        [
            {
                'key': 'KEY_ABC',
                'source': '원문',
                'target_old': '이전 번역',
                'target_new': '현재 번역',
                'status': '기존'
            },
            ...
        ]

    처리 규칙:
        1. Status == "기존"인 행만 필터링
        2. KEY 기준으로 양쪽 파일 매칭
        3. Target(D열) 값 비교
        4. 다른 것만 반환

    Reference:
        PRD v1.4.0 섹션 3.4.1
    """
    # 1. 파일1 로드 (비교1)
    wb1 = load_workbook(file1, data_only=True)
    ws1 = wb1.active

    data1 = {}  # {KEY: {'source': ..., 'target': ..., 'status': ...}}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if len(row) >= 5:
            table, key, source, target, status = row[:5]

            # Status == "기존"만 수집
            if status == "기존":
                data1[key] = {
                    'source': source,
                    'target': target,
                    'status': status
                }

    wb1.close()

    # 2. 파일2 로드 (비교2)
    wb2 = load_workbook(file2, data_only=True)
    ws2 = wb2.active

    data2 = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if len(row) >= 5:
            table, key, source, target, status = row[:5]

            # Status == "기존"만 수집
            if status == "기존":
                data2[key] = {
                    'source': source,
                    'target': target,
                    'status': status
                }

    wb2.close()

    # 3. KEY 일치 확인 (양쪽 모두 있는 KEY만 비교)
    keys1 = set(data1.keys())
    keys2 = set(data2.keys())
    common_keys = keys1 & keys2

    # KEY 불일치 정보 (로그용, 오류 발생시키지 않음)
    only_in_1 = keys1 - keys2
    only_in_2 = keys2 - keys1

    # 4. Target 비교 (다른 것만 수집)
    differences = []

    for key in sorted(common_keys):  # KEY 알파벳 순으로 처리
        target1 = data1[key]['target']
        target2 = data2[key]['target']

        # Target이 다른 경우만
        if target1 != target2:
            differences.append({
                'key': key,
                'source': data1[key]['source'],
                'target_old': target1,
                'target_new': target2,
                'status': '기존'
            })

    return differences


def create_overview_sheet(wb: Workbook, all_diffs: Dict[str, List[Dict]]) -> Dict[str, int]:
    """
    Overview 시트 생성

    Args:
        wb: Workbook 객체
        all_diffs: {언어코드: [차이 목록]}

    Returns:
        {KEY: Overview 인덱스} 매핑

    Reference:
        PRD v1.4.0 섹션 3.5.3
    """
    ws = wb.create_sheet("Overview", 0)  # 첫 번째 시트

    # 헤더
    ws.append(['#', 'KEY', 'EN', 'CT', 'CS', 'JA', 'TH', 'PT-BR', 'RU'])

    # 헤더 스타일
    header_fill = PatternFill(start_color="DBEEF4", end_color="DBEEF4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    for col in range(1, 10):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 모든 언어의 변경된 KEY 수집 (중복 제거)
    all_changed_keys = set()
    for diffs in all_diffs.values():
        for diff in diffs:
            all_changed_keys.add(diff['key'])

    # KEY 정렬 (알파벳 순)
    sorted_keys = sorted(all_changed_keys)

    # KEY -> 인덱스 매핑
    overview_key_index = {}

    # 각 KEY별로 언어별 변경 여부 확인
    for idx, key in enumerate(sorted_keys, start=1):
        row_data = [idx, key]
        overview_key_index[key] = idx

        for lang in VALID_LANGUAGES:
            # 해당 언어에서 이 KEY가 변경되었는지 확인
            lang_diffs = all_diffs.get(lang, [])
            has_change = any(d['key'] == key for d in lang_diffs)

            row_data.append('O' if has_change else 'X')

        ws.append(row_data)

    # 셀 스타일 (O/X)
    o_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    o_font = Font(name="Calibri", size=11, color="006100", bold=True)

    x_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    x_font = Font(name="Calibri", size=11, color="3C3C3C")

    center_align = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(3, 10):  # C열(3)부터 I열(9)까지
            cell = ws.cell(row_idx, col_idx)
            cell.alignment = center_align

            if cell.value == 'O':
                cell.fill = o_fill
                cell.font = o_font
            elif cell.value == 'X':
                cell.fill = x_fill
                cell.font = x_font

        # 인덱스, KEY 열 정렬
        ws.cell(row_idx, 1).alignment = center_align
        ws.cell(row_idx, 2).alignment = Alignment(horizontal="left", vertical="center")

    # 열 너비 설정
    ws.column_dimensions['A'].width = 8   # #
    ws.column_dimensions['B'].width = 50  # KEY
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 10  # 언어 코드

    # 자동 필터
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    return overview_key_index


def create_language_sheet(
    wb: Workbook,
    language_code: str,
    diffs: List[Dict],
    overview_key_index: Dict[str, int]
) -> None:
    """
    언어별 상세 시트 생성

    Args:
        wb: Workbook 객체
        language_code: 언어 코드
        diffs: 해당 언어의 차이 목록
        overview_key_index: {KEY: Overview 인덱스}

    Reference:
        PRD v1.4.0 섹션 3.5.4
    """
    ws = wb.create_sheet(language_code)

    # 헤더
    ws.append(['Overview Index', 'KEY', 'Source', '이전 Target', '현재 Target'])

    # 헤더 스타일
    header_fill = PatternFill(start_color="DBEEF4", end_color="DBEEF4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True)
    header_align = Alignment(horizontal="left", vertical="center")

    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 데이터 (Overview 인덱스 순서대로)
    sorted_diffs = sorted(diffs, key=lambda d: overview_key_index.get(d['key'], 0))

    for diff in sorted_diffs:
        ws.append([
            overview_key_index.get(diff['key'], 0),
            diff['key'],
            diff['source'],
            diff['target_old'],
            diff['target_new']
        ])

    # 데이터 행 스타일
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=11)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 6):
            cell = ws.cell(row_idx, col_idx)
            cell.font = data_font
            cell.alignment = data_align

        # 행 높이
        ws.row_dimensions[row_idx].height = 30

    # 열 너비 설정
    ws.column_dimensions['A'].width = 16  # Overview Index
    ws.column_dimensions['B'].width = 50  # KEY
    ws.column_dimensions['C'].width = 40  # Source
    ws.column_dimensions['D'].width = 40  # 이전 Target
    ws.column_dimensions['E'].width = 40  # 현재 Target

    # 자동 필터
    ws.auto_filter.ref = f"A1:E{ws.max_row}"


def legacy_diff(
    folder1: Path,
    folder2: Path,
    output_path: Path,
    progress_callback: Optional[Callable[[int, str], None]] = None
) -> Tuple[Path, Dict[str, int]]:
    """
    Legacy Diff 메인 함수

    Args:
        folder1: 비교1 폴더
        folder2: 비교2 폴더
        output_path: 출력 파일 경로
        progress_callback: 진행률 콜백

    Returns:
        (출력 파일 경로, {언어코드: 변경 개수})

    Raises:
        LegacyDiffError: 검증 실패 시

    Reference:
        PRD v1.4.0 섹션 3
    """
    if progress_callback:
        progress_callback(0, "폴더 검증 중...")

    # Step 1: 폴더 검증
    is_valid, error_msg, file_pairs = validate_diff_folders(folder1, folder2)
    if not is_valid:
        raise LegacyDiffError(error_msg, "LEGACY_DIFF_VALIDATION_ERROR")

    if progress_callback:
        progress_callback(10, "파일 비교 중...")

    # Step 2: 언어별 파일 비교
    all_diffs = {}
    total_langs = len(VALID_LANGUAGES)

    for lang_idx, lang in enumerate(VALID_LANGUAGES):
        file1, file2 = file_pairs[lang]

        diffs = compare_language_files(file1, file2, lang)
        all_diffs[lang] = diffs

        # 진행률 업데이트 (10% ~ 70%)
        if progress_callback:
            progress = 10 + int((lang_idx + 1) / total_langs * 60)
            progress_callback(progress, f"{lang} 파일 비교 중...")

    # 변경사항 없으면 오류
    total_changes = sum(len(diffs) for diffs in all_diffs.values())
    if total_changes == 0:
        raise LegacyDiffError(
            "변경된 항목이 없습니다.\n\n"
            "모든 '기존' 상태 항목의 Target이 동일합니다.",
            "LEGACY_DIFF_NO_CHANGES"
        )

    if progress_callback:
        progress_callback(75, "Overview 시트 생성 중...")

    # Step 3: 결과 파일 생성
    wb = Workbook()

    # Overview 시트 생성 (KEY -> 인덱스 매핑도 반환)
    overview_key_index = create_overview_sheet(wb, all_diffs)

    if progress_callback:
        progress_callback(85, "언어별 시트 생성 중...")

    # 언어별 시트 생성
    for lang in VALID_LANGUAGES:
        if all_diffs[lang]:  # 차이가 있는 경우만
            create_language_sheet(wb, lang, all_diffs[lang], overview_key_index)

    # 기본 시트 제거
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    if progress_callback:
        progress_callback(95, "파일 저장 중...")

    # Step 4: 파일 저장
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    if progress_callback:
        progress_callback(100, "완료")

    # 통계 정보 반환
    stats = {lang: len(diffs) for lang, diffs in all_diffs.items()}

    return output_path, stats


def generate_diff_filename() -> str:
    """
    Diff 파일명 생성

    Returns:
        YYYYMMDDHHMMSS_DIFF.xlsx 형식의 파일명
    """
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"{timestamp}_DIFF.xlsx"
