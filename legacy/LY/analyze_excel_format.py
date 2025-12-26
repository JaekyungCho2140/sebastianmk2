#!/usr/bin/env python3
"""
Excel 파일 서식 분석 스크립트
열 너비, 행 높이, 폰트, 색상, 정렬, 테두리, 고정 틀, 필터 등 추출
"""

import openpyxl
from openpyxl.styles import Font, Fill, Border, Alignment
from pathlib import Path

def analyze_excel_format(file_path):
    """Excel 파일의 서식 정보를 분석하여 출력"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    print(f"\n{'='*80}")
    print(f"파일: {Path(file_path).name}")
    print(f"시트: {ws.title}")
    print(f"{'='*80}\n")

    # 1. 열 너비
    print("## 1. 열 너비 (Column Width)")
    print("-" * 80)
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            print(f"  {col_letter}: {width:.2f} 문자 단위")
        else:
            print(f"  {col_letter}: 기본값 (약 8.43 문자)")

    # 2. 행 높이
    print("\n## 2. 행 높이 (Row Height)")
    print("-" * 80)
    for row_num in range(1, min(6, ws.max_row + 1)):  # 처음 5행만
        if row_num in ws.row_dimensions:
            height = ws.row_dimensions[row_num].height
            if height:
                print(f"  행 {row_num}: {height:.2f} 포인트")
            else:
                print(f"  행 {row_num}: 기본값 (약 15 포인트)")
        else:
            print(f"  행 {row_num}: 기본값 (약 15 포인트)")

    # 3. 헤더 행 (1행) 서식
    print("\n## 3. 헤더 행 (1행) 서식")
    print("-" * 80)
    cell_a1 = ws['A1']

    # 폰트
    font = cell_a1.font
    print(f"  폰트명: {font.name if font.name else '기본 (Calibri)'}")
    print(f"  폰트 크기: {font.size if font.size else 11}pt")
    print(f"  폰트 색상: #{font.color.rgb if font.color and hasattr(font.color, 'rgb') else '000000'}")
    print(f"  굵게: {font.bold if font.bold else False}")
    print(f"  기울임: {font.italic if font.italic else False}")

    # 배경색
    fill = cell_a1.fill
    if fill.fill_type == 'solid':
        bg_color = fill.start_color.rgb if fill.start_color and hasattr(fill.start_color, 'rgb') else 'FFFFFF'
        print(f"  배경색: #{bg_color}")
    else:
        print(f"  배경색: 없음")

    # 정렬
    alignment = cell_a1.alignment
    print(f"  수평 정렬: {alignment.horizontal if alignment.horizontal else '일반'}")
    print(f"  수직 정렬: {alignment.vertical if alignment.vertical else 'bottom'}")
    print(f"  텍스트 줄바꿈: {alignment.wrap_text if alignment.wrap_text else False}")

    # 테두리
    border = cell_a1.border
    print(f"  테두리:")
    print(f"    상단: {border.top.style if border.top.style else '없음'}")
    print(f"    하단: {border.bottom.style if border.bottom.style else '없음'}")
    print(f"    좌측: {border.left.style if border.left.style else '없음'}")
    print(f"    우측: {border.right.style if border.right.style else '없음'}")

    # 4. 데이터 행 (2행) 서식
    print("\n## 4. 데이터 행 (2행) 서식")
    print("-" * 80)
    cell_a2 = ws['A2']

    # 폰트
    font2 = cell_a2.font
    print(f"  폰트명: {font2.name if font2.name else '기본 (Calibri)'}")
    print(f"  폰트 크기: {font2.size if font2.size else 11}pt")
    print(f"  폰트 색상: #{font2.color.rgb if font2.color and hasattr(font2.color, 'rgb') else '000000'}")
    print(f"  굵게: {font2.bold if font2.bold else False}")

    # 배경색
    fill2 = cell_a2.fill
    if fill2.fill_type == 'solid':
        bg_color2 = fill2.start_color.rgb if fill2.start_color and hasattr(fill2.start_color, 'rgb') else 'FFFFFF'
        print(f"  배경색: #{bg_color2}")
    else:
        print(f"  배경색: 없음")

    # 정렬
    alignment2 = cell_a2.alignment
    print(f"  수평 정렬: {alignment2.horizontal if alignment2.horizontal else '일반'}")
    print(f"  수직 정렬: {alignment2.vertical if alignment2.vertical else 'bottom'}")
    print(f"  텍스트 줄바꿈: {alignment2.wrap_text if alignment2.wrap_text else False}")

    # 5. 고정 틀 (Freeze Panes)
    print("\n## 5. 고정 틀 (Freeze Panes)")
    print("-" * 80)
    if ws.freeze_panes:
        print(f"  고정 기준 셀: {ws.freeze_panes}")
        if ws.freeze_panes == 'A2':
            print(f"  → 첫 번째 행(헤더) 고정")
        else:
            print(f"  → 사용자 정의 고정")
    else:
        print(f"  없음")

    # 6. 자동 필터 (Auto Filter)
    print("\n## 6. 자동 필터 (Auto Filter)")
    print("-" * 80)
    if ws.auto_filter.ref:
        print(f"  범위: {ws.auto_filter.ref}")
        print(f"  → 필터 활성화됨")
    else:
        print(f"  없음")

    wb.close()
    print("\n" + "="*80 + "\n")

if __name__ == "__main__":
    # 병합 파일 분석
    merged_file = r"D:\Repository\LY_Table\.claude\docs\Merged Table\251104_LYGL_StringALL.xlsx"
    analyze_excel_format(merged_file)

    # 언어별 파일 분석 (EN 예시)
    lang_file = r"D:\Repository\LY_Table\.claude\docs\Tables by Language\251104_EN.xlsx"
    analyze_excel_format(lang_file)
