import pandas as pd
import datetime
import tkinter as tk
from tkinter import messagebox, filedialog
import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import stat
import threading
import queue
from openpyxl.worksheet.worksheet import Worksheet
from progress_window import ProgressWindow  # 새로 만든 모듈 임포트
from PIL import Image, ImageTk
import cProfile
import pstats
from io import StringIO
import multiprocessing as mp
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor

class ImageButton(tk.Label):
    def __init__(self, parent, normal_image, hover_image, command=None, **kwargs):
        self.normal_image = ImageTk.PhotoImage(Image.open(normal_image))
        self.hover_image = ImageTk.PhotoImage(Image.open(hover_image))
        super().__init__(parent, image=self.normal_image, bd=0, highlightthickness=0, **kwargs)  # 테두리 제거
        
        self.command = command
        
        self.bind('<Enter>', self._on_enter)
        self.bind('<Leave>', self._on_leave)
        self.bind('<Button-1>', self._on_click)
        
    def _on_enter(self, event):
        self.configure(image=self.hover_image)
        
    def _on_leave(self, event):
        self.configure(image=self.normal_image)
        
    def _on_click(self, event):
        if self.command:
            self.command()

def center_window(window, root):
    """주어진 창을 화면 중앙에 배치합니다."""
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (root.winfo_x() + (root.winfo_width() // 2)) - (width // 2)
    y = (root.winfo_y() + (root.winfo_height() // 2)) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')

def profile_function(func):
    def wrapper(*args, **kwargs):
        pr = cProfile.Profile()
        pr.enable()
        result = func(*args, **kwargs)
        pr.disable()
        s = StringIO()
        ps = pstats.Stats(pr, stream=s).sort_stats('cumulative')
        ps.print_stats()
        print(s.getvalue())
        return result
    return wrapper

@profile_function
def read_excel_file(file_path, sheet_name, header_row, skip_rows):
    start_time = time.time()
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, skiprows=skip_rows)
    end_time = time.time()
    print(f"파일 읽기 시간 ({os.path.basename(file_path)}): {end_time - start_time:.2f}초")
    return df

# 데이터 읽기 전에 먼저 파일이 존재하는지 확인하고, 열이 존재하는지 확인
def run_merge(progress_queue, folder_path, start_time):
    try:
        # 파일 경로 설정
        cinematic_path = os.path.join(folder_path, "CINEMATIC_DIALOGUE.xlsm")
        smalltalk_path = os.path.join(folder_path, "SMALLTALK_DIALOGUE.xlsm")
        npc_path = os.path.join(folder_path, "NPC.xlsm")

        # 파일 존재 여부 확인
        missing_files = []
        if not os.path.isfile(cinematic_path):
            missing_files.append(f"파일을 찾을 수 없습니다: {cinematic_path}")
        if not os.path.isfile(smalltalk_path):
            missing_files.append(f"파일을 찾을 수 없습니다: {smalltalk_path}")
        if not os.path.isfile(npc_path):
            missing_files.append(f"파일을 찾을 수 없습니다: {npc_path}")
        if missing_files:
            raise FileNotFoundError("\n".join(missing_files))

        # 단계 정보 전송
        progress_queue.put("단계:1/3")
        progress_queue.put("파일:CINEMATIC_DIALOGUE.xlsm")

        # 데이터 읽기
        cinematic_data = read_excel_file(cinematic_path, sheet_name=1, header_row=1, skip_rows=9)
        progress_queue.put(20)
        progress_queue.put("처리된 파일:1")

        progress_queue.put("파일:SMALLTALK_DIALOGUE.xlsm")
        smalltalk_data = read_excel_file(smalltalk_path, sheet_name=1, header_row=1, skip_rows=4)
        progress_queue.put(40)
        progress_queue.put("처리된 파일:2")

        # 단계 업데이트
        progress_queue.put("단계:2/3")
        progress_queue.put("파일:데이터 병합 중...")

        # 열 인덱스가 범위를 벗어나지 않는지 확인
        cinematic_cols = cinematic_data.shape[1]
        smalltalk_cols = smalltalk_data.shape[1]
        
        # 결과물 파일의 헤더 설정
        headers = ['#', 'Table Name', 'String ID', 'Table/ID', 'NPC ID', 'Speaker Name',
                   'KO (M)', 'KO (F)', 'EN (M)', 'EN (F)', 'CT (M)', 'CT (F)', 'CS (M)',
                   'CS (F)', 'JA (M)', 'JA (F)', 'TH (M)', 'TH (F)', 'ES-LATAM (M)', 'ES-LATAM (F)',
                   'PT-BR (M)', 'PT-BR (F)', 'NOTE']

        # 결과 데이터 프레임 생성 - 미리 열을 생성해둠
        result_df = pd.DataFrame(columns=headers)
        
        # 각 데이터프레임의 길이 확인
        cin_len = len(cinematic_data)
        small_len = len(smalltalk_data)
        total_len = cin_len + small_len
        
        # 결과 데이터프레임에 필요한 개수만큼 행 추가 (빈 행으로)
        result_df = pd.DataFrame(index=range(total_len), columns=headers)
        
        # 인덱스 열 채우기
        result_df['#'] = range(1, total_len + 1)
        
        # Table Name 열 채우기
        result_df.loc[:cin_len-1, 'Table Name'] = 'CINEMATIC_DIALOGUE'
        result_df.loc[cin_len:, 'Table Name'] = 'SMALLTALK_DIALOGUE'
        
        # 안전하게 열 인덱스 확인하고 데이터 할당
        # String ID 열 (인덱스 7)
        if 7 < cinematic_cols and 7 < smalltalk_cols:
            result_df.loc[:cin_len-1, 'String ID'] = cinematic_data.iloc[:, 7].values
            result_df.loc[cin_len:, 'String ID'] = smalltalk_data.iloc[:, 7].values
        
        # Table/ID 열 생성
        result_df['Table/ID'] = result_df['Table Name'] + '/' + result_df['String ID'].astype(str)
        
        # NPC ID 열 (인덱스 8)
        if 8 < cinematic_cols and 8 < smalltalk_cols:
            result_df.loc[:cin_len-1, 'NPC ID'] = cinematic_data.iloc[:, 8].values
            result_df.loc[cin_len:, 'NPC ID'] = smalltalk_data.iloc[:, 8].values
        
        # 나머지 언어 데이터 열 매핑
        language_mapping = {
            'KO (M)': (11, 12),
            'KO (F)': (12, 13),
            'EN (M)': (13, 14),
            'EN (F)': (14, 15),
            'CT (M)': (15, 16),
            'CT (F)': (16, 17),
            'CS (M)': (17, 18),
            'CS (F)': (18, 19),
            'JA (M)': (19, 20),
            'JA (F)': (20, 21),
            'TH (M)': (21, 22),
            'TH (F)': (22, 23),
            'ES-LATAM (M)': (23, 24),
            'ES-LATAM (F)': (24, 25),
            'PT-BR (M)': (25, 26),
            'PT-BR (F)': (26, 27),
            'NOTE': (29, 30)
        }
        
        # 각 언어 열에 데이터 안전하게 채우기
        for col_name, (cin_idx, small_idx) in language_mapping.items():
            # 인덱스가 범위 내에 있는지 확인
            if cin_idx < cinematic_cols and small_idx < smalltalk_cols:
                result_df.loc[:cin_len-1, col_name] = cinematic_data.iloc[:, cin_idx].values
                result_df.loc[cin_len:, col_name] = smalltalk_data.iloc[:, small_idx].values
            else:
                # 인덱스가 범위를 벗어나면 빈 값으로 설정
                result_df[col_name] = ''
        
        progress_queue.put(60)
        progress_queue.put("단계:3/3")
        progress_queue.put("파일:NPC.xlsm")
        
        # 원본 3(NPC.xlsm) 데이터 읽기 (두 번째 시트)
        npc_data = pd.read_excel(npc_path, sheet_name='NPC', header=1)
        npc_data = npc_data.drop_duplicates(subset=npc_data.columns[7])
        progress_queue.put("처리된 파일:3")
        
        # 결과 파일의 'NPC ID' 열과 원본 3의 'H열 유니크 아이디' 열을 기준으로 'J열 NPC 이름' 값을 불러오기
        # 안전하게 매핑을 위해 try-except 구문 사용
        try:
            # Dictionary 매핑을 생성
            npc_map = dict(zip(npc_data.iloc[:, 7], npc_data.iloc[:, 9]))
            # 매핑된 값이 없으면 원래 값을 유지
            result_df['Speaker Name'] = result_df['NPC ID'].map(npc_map).fillna(result_df['NPC ID'])
        except Exception as e:
            # 매핑 실패시 오류 메시지 표시
            progress_queue.put(("error", f"NPC 이름 매핑 중 오류 발생: {str(e)}"))
            return
        
        # 9번째 열이 빈 셀(NaN) 또는 0 또는 '미사용'인 행 제거
        # 'EN (M)' 열이 9번째 열에 해당
        result_df = result_df[~(pd.isna(result_df['EN (M)']) | result_df['EN (M)'].isin([0, '미사용']))]
        
        # 인덱스 열 갱신
        result_df['#'] = range(1, len(result_df) + 1)
        
        progress_queue.put(80)
        progress_queue.put("파일:결과 파일 저장 중...")
        
        # 출력 파일 이름 설정
        date_str = datetime.datetime.now().strftime('%m%d')
        output_file = f'{date_str}_MIR4_MASTER_DIALOGUE.xlsx'

        # 파일이 이미 존재할 경우 다른 이름으로 저장
        counter = 1
        while os.path.exists(output_file):
            output_file = f'{date_str}_MIR4_MASTER_DIALOGUE_{counter}.xlsx'
            counter += 1

        # 결과 파일 저장 (엑셀)
        result_df.to_excel(output_file, index=False)

        # 결과 파일 서식 지정
        wb = load_workbook(output_file)
        ws = wb.active

        # 폰트 및 서식 설정
        header_font = Font(name='맑은 고딕', size=12, bold=True, color='9C5700')
        default_font = Font(name='맑은 고딕', size=10)
        header_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        border_style = Side(border_style='thin', color='000000')
        full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        # 헤더 행 서식 지정
        if isinstance(ws, Worksheet):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = full_border

            # 나머지 셀 서식 지정
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = default_font
                    cell.border = full_border

            # 틀 고정
            ws.freeze_panes = 'A2'

        # 서식 지정된 파일 저장
        wb.save(output_file)

        # 결과 파일 읽기 전용 설정
        os.chmod(output_file, stat.S_IREAD)

        progress_queue.put(100)

        elapsed_time = time.time() - start_time
        progress_queue.put(f"완료:파일이 {output_file}로 저장되었습니다. 소요 시간: {int(elapsed_time)}초")

    except Exception as e:
        progress_queue.put(("error", str(e)))

def run_merge_string(progress_queue, folder_path, start_time):
    try:
        # 파일 경로 설정
        file_list = [
            "SEQUENCE_DIALOGUE.xlsm",
            "STRING_BUILTIN.xlsm",
            "STRING_MAIL.xlsm",
            "STRING_MESSAGE.xlsm",
            "STRING_NPC.xlsm",
            "STRING_QUESTTEMPLATE.xlsm",
            "STRING_TEMPLATE.xlsm",
            "STRING_TOOLTIP.xlsm"
        ]

        # 각 파일의 헤더와 데이터 시작 행
        header_rows = {
            "SEQUENCE_DIALOGUE.xlsm": 2,
            "STRING_BUILTIN.xlsm": 2,
            "STRING_MAIL.xlsm": 2,
            "STRING_MESSAGE.xlsm": 2,
            "STRING_NPC.xlsm": 2,
            "STRING_QUESTTEMPLATE.xlsm": 2,
            "STRING_TEMPLATE.xlsm": 2,
            "STRING_TOOLTIP.xlsm": 2
        }

        start_rows = {
            "SEQUENCE_DIALOGUE.xlsm": 9,
            "STRING_BUILTIN.xlsm": 4,
            "STRING_MAIL.xlsm": 4,
            "STRING_MESSAGE.xlsm": 4,
            "STRING_NPC.xlsm": 4,
            "STRING_QUESTTEMPLATE.xlsm": 7,
            "STRING_TEMPLATE.xlsm": 4,
            "STRING_TOOLTIP.xlsm": 4
        }

        # 매칭되는 열 인덱스 설정
        matching_columns = {
            "SEQUENCE_DIALOGUE.xlsm": [7, None, 10, 11, 12, 13, 14, 15, 16, 17, None, None],
            "STRING_BUILTIN.xlsm": [7, 21, 8, 9, 10, 11, 12, 13, 14, 15, None, None],
            "STRING_MAIL.xlsm": [7, None, 8, 9, 10, 11, 12, 13, 14, 15, None, None],
            "STRING_MESSAGE.xlsm": [7, 21, 8, 9, 10, 11, 12, 13, 14, 15, None, None],
            "STRING_NPC.xlsm": [7, 20, 9, 10, 11, 12, 13, 14, 15, 16, 18, 19],
            "STRING_QUESTTEMPLATE.xlsm": [7, 0, 12, 13, 14, 15, 16, 17, 18, 19, None, None],
            "STRING_TEMPLATE.xlsm": [7, 19, 8, 9, 10, 11, 12, 13, 14, 15, None, 18],
            "STRING_TOOLTIP.xlsm": [7, 8, 11, 12, 13, 14, 15, 16, 17, 18, None, None]
        }

        # 결과물 파일의 헤더 설정
        headers = ['#', 'Table Name', 'String ID', 'Table/ID', 'NOTE', 'KO', 'EN', 'CT', 'CS', 'JA', 'TH', 'ES-LATAM', 'PT-BR', 'NPC 이름', '비고']

        # 결과 데이터 프레임 생성
        result_df = pd.DataFrame(columns=headers)

        # 단계 정보 전송
        progress_queue.put("단계:1/2")
        progress_queue.put("파일:파일 읽는 중...")

        # 각 파일에서 데이터 읽어오기
        for i, file in enumerate(file_list):
            file_path = os.path.join(folder_path, file)
            if not os.path.isfile(file_path):
                progress_queue.put(("error", f"파일을 찾을 수 없습니다: {file_path}"))
                return
            
            progress_queue.put(f"파일:{file}")
            data = read_excel_file(file_path, sheet_name=1, header_row=header_rows[file], skip_rows=start_rows[file])

            # Table Name 열 채우기
            table_name = file.replace(".xlsm", "")
            temp_df = pd.DataFrame({
                '#': range(len(result_df) + 1, len(result_df) + len(data) + 1),
                'Table Name': table_name,
                'String ID': data.iloc[:, matching_columns[file][0]] if matching_columns[file][0] is not None else '',
                'Table/ID': table_name + '/' + data.iloc[:, matching_columns[file][0]].astype(str) if matching_columns[file][0] is not None else '',
                'NOTE': data.iloc[:, matching_columns[file][1]] if matching_columns[file][1] is not None else '',
                'KO': data.iloc[:, matching_columns[file][2]] if matching_columns[file][2] is not None else '',
                'EN': data.iloc[:, matching_columns[file][3]] if matching_columns[file][3] is not None else '',
                'CT': data.iloc[:, matching_columns[file][4]] if matching_columns[file][4] is not None else '',
                'CS': data.iloc[:, matching_columns[file][5]] if matching_columns[file][5] is not None else '',
                'JA': data.iloc[:, matching_columns[file][6]] if matching_columns[file][6] is not None else '',
                'TH': data.iloc[:, matching_columns[file][7]] if matching_columns[file][7] is not None else '',
                'ES-LATAM': data.iloc[:, matching_columns[file][8]] if matching_columns[file][8] is not None else '',
                'PT-BR': data.iloc[:, matching_columns[file][9]] if matching_columns[file][9] is not None else '',
                'NPC 이름': data.iloc[:, matching_columns[file][10]] if matching_columns[file][10] is not None else '',
                '비고': data.iloc[:, matching_columns[file][11]] if matching_columns[file][11] is not None else ''
            })
            result_df = pd.concat([result_df, temp_df], ignore_index=True)

            progress_queue.put(int(20 + (50 / len(file_list)) * (i + 1)))
            progress_queue.put(f"처리된 파일:{i+1}")

        progress_queue.put("단계:2/2")
        progress_queue.put("파일:결과 파일 저장 중...")

        # 7번째 열(인덱스 6)이 빈 셀(NaN) 또는 0 또는 '미사용'인 행 제거
        result_df = result_df[~(pd.isna(result_df.iloc[:, 6]) | result_df.iloc[:, 6].isin([0, '미사용']))]

        # 인덱스 열 갱신
        result_df['#'] = range(1, len(result_df) + 1)

        # 출력 파일 이름 설정
        date_str = datetime.datetime.now().strftime('%m%d')
        output_file = f'{date_str}_MIR4_MASTER_STRING.xlsx'

        # 파일이 이미 존재할 경우 다른 이름으로 저장
        counter = 1
        while os.path.exists(output_file):
            output_file = f'{date_str}_MIR4_MASTER_STRING_{counter}.xlsx'
            counter += 1

        # 결과 파일 저장 (엑셀)
        result_df.to_excel(output_file, index=False)

        # 결과 파일 서식 지정
        wb = load_workbook(output_file)
        ws = wb.active

        # 폰트 및 서식 설정
        header_font = Font(name='맑은 고딕', size=12, bold=True, color='9C5700')
        default_font = Font(name='맑은 고딕', size=10)
        header_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        border_style = Side(border_style='thin', color='000000')
        full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        # 헤더 행 서식 지정
        if isinstance(ws, Worksheet):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = full_border

            # 나머지 셀 서식 지정
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = default_font
                    cell.border = full_border

            # 틀 고정
            ws.freeze_panes = 'A2'

        # 서식 지정된 파일 저장
        wb.save(output_file)

        # 결과 파일 읽기 전용 설정
        os.chmod(output_file, stat.S_IREAD)

        progress_queue.put(100)

        elapsed_time = time.time() - start_time
        progress_queue.put(f"완료:파일이 {output_file}로 저장되었습니다. 소요 시간: {int(elapsed_time)}초")

    except Exception as e:
        progress_queue.put(("error", str(e)))

def start_merge():
    try:
        # 경로 선택 대화 상자 열기
        folder_path = filedialog.askdirectory(title="대화 파일이 포함된 폴더 선택")
        if not folder_path:
            return  # 경로가 선택되지 않으면 함수 종료

        start_time = time.time()
        
        # 진행 창 생성 (모듈화된 클래스 사용)
        progress_window = ProgressWindow(
            root, 
            title="M4 DIALOGUE 병합 중",
            theme_color="#4CAF50",
            font_family="맑은 고딕"
        )
        
        # 진행 작업 시작
        progress_window.start(
            run_merge, 
            args=(folder_path, start_time),
            total_steps=3,
            total_files=3
        )

    except Exception as e:
        messagebox.showerror("오류", str(e), parent=root)

def start_merge_string():
    try:
        # 경로 선택 대화 상자 열기
        folder_path = filedialog.askdirectory(title="대화 파일이 포함된 폴더 선택")
        if not folder_path:
            return  # 경로가 선택되지 않으면 함수 종료

        start_time = time.time()
        
        # 진행 창 생성 (모듈화된 클래스 사용)
        progress_window = ProgressWindow(
            root, 
            title="M4 STRING 병합 중",
            theme_color="#2196F3",  # STRING은 파란색 테마
            font_family="맑은 고딕"
        )
        
        # 진행 작업 시작
        progress_window.start(
            run_merge_string, 
            args=(folder_path, start_time),
            total_steps=2,
            total_files=8
        )

    except Exception as e:
        messagebox.showerror("오류", str(e), parent=root)

# 루트 창 생성
root = tk.Tk()
root.title("M4/GL Merged")
root.geometry("480x640")
root.resizable(False, False)  # 창 크기 고정

# 배경 이미지 설정
bg_image = ImageTk.PhotoImage(Image.open("cho_MM_480x640.jpg"))
bg_label = tk.Label(root, image=bg_image)
bg_label.place(x=0, y=0)

# DIALOGUE 버튼 생성
dialogue_button = ImageButton(
    root,
    "cho_MIR4_button_D_82x30.jpg",
    "cho_MIR4_button_D_click_82x30.jpg",
    command=start_merge
)
dialogue_button.place(x=199, y=348)

# STRING 버튼 생성
string_button = ImageButton(
    root,
    "cho_MIR4_button_S_56x30.jpg",
    "cho_MIR4_button_S_click_56x30.jpg",
    command=start_merge_string
)
string_button.place(x=212, y=388)

# 종료 버튼 생성
exit_button = ImageButton(
    root,
    "cho_MIR4_button_E_50x25.jpg",
    "cho_MIR4_button_E_click_50x25.jpg",
    command=root.quit
)
exit_button.place(x=315, y=505)

# 이벤트 루프 실행
root.mainloop()
